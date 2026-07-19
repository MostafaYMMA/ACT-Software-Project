import csv
import os
import sqlite3
import sys
import tempfile
import unittest
from datetime import date, datetime
from unittest.mock import patch

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "services"))

import storage_service
import date_utils


class TempDbTestCase(unittest.TestCase):
    """Shared scaffolding: every test runs against a throwaway database,
    never the real data/cards.db."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.db_path = os.path.join(self.temp_dir.name, "cards_test.db")

        self.db_path_patcher = patch.object(storage_service, "DB_PATH", self.db_path)
        self.db_path_patcher.start()
        self.addCleanup(self.db_path_patcher.stop)

        storage_service.init_db()

    def _insert_row(self, table, *, day="Monday, 06 Jul", subject="S",
                    received="2026-07-06 10:00:00", project_number="P1",
                    project_name="FBGIU Cafe", task="T1", qty="8",
                    rate=0.0, period="2026-07-04 to 2026-07-10",
                    person_number="100", sender="s@x.com"):
        conn = sqlite3.connect(self.db_path)
        conn.execute(
            f'INSERT INTO "{table}" '
            f'(day, "Date", subject, sender, "Project Number", "Project Name", '
            f'"Task Name", "Qty", rate, period, person_number, received, received_month) '
            f'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (day, received[:10], subject, sender, project_number, project_name,
             task, qty, rate, period, person_number, received, received[:7]),
        )
        conn.commit()
        conn.close()

    def _export_path(self, name="export.xlsx"):
        return os.path.join(self.temp_dir.name, name)

    def _db_rows(self, sql, params=()):
        conn = sqlite3.connect(self.db_path)
        try:
            return conn.execute(sql, params).fetchall()
        finally:
            conn.close()


class LastMonthRangeTests(unittest.TestCase):
    """The History page's "Export Last Month" button computes its range with
    _last_month_range -- the whole calendar month before today."""

    @classmethod
    def setUpClass(cls):
        from ui.Pages.History import _last_month_range
        cls.last_month_range = staticmethod(_last_month_range)

    def test_mid_year_month(self):
        start, end = self.last_month_range(date(2026, 7, 19))
        self.assertEqual(start, date(2026, 6, 1))
        self.assertEqual(end, date(2026, 6, 30))

    def test_january_wraps_to_december_of_previous_year(self):
        start, end = self.last_month_range(date(2026, 1, 5))
        self.assertEqual(start, date(2025, 12, 1))
        self.assertEqual(end, date(2025, 12, 31))

    def test_march_in_leap_year_covers_feb_29(self):
        start, end = self.last_month_range(date(2024, 3, 10))
        self.assertEqual(start, date(2024, 2, 1))
        self.assertEqual(end, date(2024, 2, 29))

    def test_first_day_of_month_still_means_previous_month(self):
        start, end = self.last_month_range(date(2026, 8, 1))
        self.assertEqual(start, date(2026, 7, 1))
        self.assertEqual(end, date(2026, 7, 31))


class ExportHistoryLogTests(TempDbTestCase):
    """The log shown in the page's table (get_export_history) and the
    "From last export" button's backing value (get_last_export_date)."""

    def test_log_empty_before_any_export(self):
        self.assertEqual(storage_service.get_export_history(), [])

    def test_last_export_date_none_before_any_export(self):
        self.assertIsNone(storage_service.get_last_export_date())

    def test_log_returns_newest_first(self):
        conn = sqlite3.connect(self.db_path)
        conn.executemany(
            "INSERT INTO export_history (name, date) VALUES (?, ?)",
            [("old.xlsx", "2026-06-01 09:00:00"), ("new.xlsx", "2026-07-01 09:00:00")],
        )
        conn.commit()
        conn.close()
        rows = storage_service.get_export_history()
        self.assertEqual([name for name, _ in rows], ["new.xlsx", "old.xlsx"])

    def test_log_breaks_timestamp_ties_by_insertion_order(self):
        conn = sqlite3.connect(self.db_path)
        conn.executemany(
            "INSERT INTO export_history (name, date) VALUES (?, ?)",
            [("first.xlsx", "2026-07-01 09:00:00"), ("second.xlsx", "2026-07-01 09:00:00")],
        )
        conn.commit()
        conn.close()
        rows = storage_service.get_export_history()
        self.assertEqual([name for name, _ in rows], ["second.xlsx", "first.xlsx"])

    def test_export_appends_basename_to_log(self):
        self._insert_row("timecards_approved")
        path = self._export_path("timecards_2026-07-01_to_2026-07-31.xlsx")
        storage_service.export_act_invoice_overview_range("2026-07-01", "2026-07-31", path)
        rows = storage_service.get_export_history()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "timecards_2026-07-01_to_2026-07-31.xlsx")

    def test_last_export_date_tracks_range_end_and_is_overwritten(self):
        self._insert_row("timecards_approved")
        storage_service.export_act_invoice_overview_range(
            "2026-07-01", "2026-07-15", self._export_path("a.xlsx"))
        self.assertEqual(storage_service.get_last_export_date(), "2026-07-15")
        storage_service.export_act_invoice_overview_range(
            "2026-07-15", "2026-07-31", self._export_path("b.xlsx"))
        self.assertEqual(storage_service.get_last_export_date(), "2026-07-31")


class ActInvoiceExportFilterTests(TempDbTestCase):
    """What goes INTO the file: approved rows only, within the date range,
    optionally narrowed to one division -- the same choices the page's
    project-type toggle and date pickers offer."""

    def setUp(self):
        super().setUp()
        # In range (July 2026):
        self._insert_row("timecards_approved", subject="FB-A", project_number="FB1",
                         project_name="FBGIU Cafe", received="2026-07-05 10:00:00")
        self._insert_row("timecards_approved", subject="FB-B-lower", project_number="FB2",
                         project_name=" fbgiu Bakery", day="Tuesday, 07 Jul",
                         received="2026-07-07 10:00:00")
        self._insert_row("timecards_approved", subject="HL-A", project_number="HL1",
                         project_name="HLGIU Hotel", day="Wednesday, 08 Jul",
                         received="2026-07-08 10:00:00")
        self._insert_row("timecards_approved", subject="NoDiv", project_number="GX1",
                         project_name="Global Support Upgrade", day="Thursday, 09 Jul",
                         received="2026-07-09 10:00:00")
        # In range but not approved:
        self._insert_row("timecards_pending", subject="Pend", project_number="FB3",
                         received="2026-07-10 10:00:00")
        self._insert_row("timecards_rejected", subject="Rej", project_number="FB4",
                         received="2026-07-11 10:00:00")
        # Approved but outside the range:
        self._insert_row("timecards_approved", subject="June", project_number="FB5",
                         day="Friday, 26 Jun", received="2026-06-26 10:00:00")

    def _export(self, project_type=None, start="2026-07-01", end="2026-07-31"):
        path = self._export_path()
        count = storage_service.export_act_invoice_overview_range(
            start, end, path, project_type=project_type)
        return count, path

    def _labor_project_numbers(self, path):
        ws = load_workbook(path)["Billed Invoice Details"]
        numbers = []
        for row in ws.iter_rows(min_row=5):
            if row[9].value == 1:  # column J: Line == 1 marks a LABOR row
                numbers.append(row[11].value)  # column L: Project Number
        return numbers

    def test_exports_only_approved_rows_in_range(self):
        count, path = self._export()
        self.assertEqual(count, 4)
        self.assertEqual(sorted(self._labor_project_numbers(path)),
                         ["FB1", "FB2", "GX1", "HL1"])

    def test_beverage_narrows_to_fb_projects_case_insensitively(self):
        count, path = self._export(project_type="beverage")
        self.assertEqual(count, 2)
        self.assertEqual(sorted(self._labor_project_numbers(path)), ["FB1", "FB2"])

    def test_hospitality_narrows_to_hl_projects(self):
        count, path = self._export(project_type="hospitality")
        self.assertEqual(count, 1)
        self.assertEqual(self._labor_project_numbers(path), ["HL1"])

    def test_range_bounds_are_inclusive(self):
        count, _ = self._export(start="2026-07-05", end="2026-07-08")
        self.assertEqual(count, 3)  # FB-A on the 5th and HL-A on the 8th both included

    def test_flags_only_exported_rows_as_exported(self):
        self._export(project_type="beverage")
        flagged = dict(self._db_rows(
            "SELECT subject, is_exported FROM timecards_approved"))
        self.assertEqual(flagged["FB-A"], 1)
        self.assertEqual(flagged["FB-B-lower"], 1)
        self.assertEqual(flagged["HL-A"], 0)
        self.assertEqual(flagged["NoDiv"], 0)
        self.assertEqual(flagged["June"], 0)

    def test_pending_and_rejected_rows_never_flagged(self):
        self._export()
        self.assertEqual(self._db_rows("SELECT is_exported FROM timecards_pending"), [(0,)])
        self.assertEqual(self._db_rows("SELECT is_exported FROM timecards_rejected"), [(0,)])

    def test_empty_range_returns_zero_but_still_writes_file_and_log(self):
        count, path = self._export(start="2025-01-01", end="2025-01-31")
        self.assertEqual(count, 0)
        self.assertTrue(os.path.exists(path))
        ws = load_workbook(path)["Billed Invoice Details"]
        self.assertEqual(ws.cell(row=4, column=2).value, "Date")
        # Current behavior: even an empty export is logged and moves the
        # last-export marker.
        self.assertEqual(len(storage_service.get_export_history()), 1)
        self.assertEqual(storage_service.get_last_export_date(), "2025-01-31")


class ActInvoiceWorkbookLayoutTests(TempDbTestCase):
    """The produced .xlsx must match the ACT "Invoice Overview per Period"
    template: title, header row, LABOR+Expense row pairs, totals block."""

    def setUp(self):
        super().setUp()
        self._insert_row(
            "timecards_approved", subject="Only", project_number="FB1",
            project_name="FBGIU Cafe", task="Bar setup", qty="7.5", rate=150.0,
            day="Monday, 06 Jul", period="2026-07-04 to 2026-07-10",
            received="2026-07-06 10:00:00",
        )
        self.path = self._export_path()
        self.count = storage_service.export_act_invoice_overview_range(
            "2026-07-01", "2026-07-31", self.path)
        self.ws = load_workbook(self.path)["Billed Invoice Details"]

    def test_title_and_headers(self):
        self.assertEqual(self.ws["B2"].value, storage_service._ACT_TITLE)
        headers = [self.ws.cell(row=4, column=c).value for c in range(2, 17)]
        self.assertEqual(headers, storage_service._ACT_HEADERS)

    def test_labor_row_values_sit_under_matching_headers(self):
        header_for = {c: self.ws.cell(row=4, column=c).value for c in range(2, 17)}
        labor = {header_for[c]: self.ws.cell(row=5, column=c).value for c in range(2, 17)}
        self.assertEqual(labor["Date"], "Monday, 06 Jul")
        self.assertEqual(labor["Project Name"], "FBGIU Cafe")
        self.assertEqual(labor["Period"], "2026-07-04 to 2026-07-10")
        self.assertEqual(labor["Task Name"], "Bar setup")
        self.assertEqual(labor["PO"], storage_service._ACT_PO)
        self.assertEqual(labor["Line"], 1)
        self.assertEqual(labor["Type"], "LABOR")
        self.assertEqual(labor["Project Number"], "FB1")
        self.assertEqual(labor["Qty"], 7.5)
        self.assertEqual(labor["Sales Price"], 150.0)
        self.assertEqual(labor["Total Amount"], "=N5*O5")

    def test_each_timecard_gets_a_companion_expense_row(self):
        self.assertEqual(self.ws.cell(row=6, column=10).value, 2)         # Line
        self.assertEqual(self.ws.cell(row=6, column=11).value, "Expense")  # Type
        self.assertEqual(self.ws.cell(row=6, column=15).value, 0)          # Sales Price
        self.assertEqual(self.ws.cell(row=6, column=16).value, "=N6*O6")   # Total Amount

    def test_money_cells_use_aed_format(self):
        for cell in (self.ws.cell(row=5, column=15), self.ws.cell(row=5, column=16)):
            self.assertIn("AED", cell.number_format)

    def test_totals_block_follows_data(self):
        # One source row -> data rows 5-6, totals start at row 7.
        self.assertEqual(self.ws.cell(row=7, column=12).value, "Total LABOR")
        self.assertEqual(self.ws.cell(row=7, column=14).value, "=SUBTOTAL(109,Table4[Qty])")
        self.assertEqual(self.ws.cell(row=7, column=16).value,
                         '=SUMIF(Table4[Type],"LABOR",Table4[Total Amount])')
        self.assertEqual(self.ws.cell(row=8, column=12).value, "Total EXPENSE")
        self.assertEqual(self.ws.cell(row=8, column=16).value,
                         '=SUMIF(Table4[Type],"Expense",Table4[Total Amount])')
        self.assertEqual(self.ws.cell(row=9, column=12).value, "Invoice Total ")
        self.assertEqual(self.ws.cell(row=9, column=16).value, "=P7+P8")
        self.assertEqual(self.ws.cell(row=10, column=13).value, "VAT")
        self.assertEqual(self.ws.cell(row=10, column=16).value, "=P9*0.05")
        self.assertEqual(self.ws.cell(row=11, column=13).value, "Total with VAT")
        self.assertEqual(self.ws.cell(row=11, column=16).value, "=P9+P10")

    def test_returns_source_row_count_not_spreadsheet_row_count(self):
        self.assertEqual(self.count, 1)


class DashboardAlignmentTests(TempDbTestCase):
    """The page's status text says "Exported N row(s) (received {start} to
    {end})" -- N must be the same number the Dashboard reports as Approved
    for that window and division. Rows are written through save_cards, the
    same path a real Outlook scan uses, so "Date" is derived from
    "received" exactly as in production."""

    def setUp(self):
        super().setUp()

        def entry(subject, project_code, project_name, received, day, hours, person):
            return {
                "day": day, "received": received, "labor_type": "Regular",
                "time_type": "Worked", "hours": hours, "project_code": project_code,
                "project_name": project_name, "task": "T1", "name": "Someone",
                "period": "2026-07-04 to 2026-07-10", "person_number": person,
                "subject": subject, "sender": "s@x.com", "status": "Approved",
            }

        storage_service.save_cards([
            entry("FB-A", "FB1", "FBGIU Cafe", "2026-07-05 10:00:00", "Monday, 06 Jul", "8", "100"),
            entry("HL-A", "HL1", "HLGIU Hotel", "2026-07-08 10:00:00", "Tuesday, 07 Jul", "4", "101"),
            # Received late on the range's last day: the export's string
            # comparison and the Dashboard's end-of-day extension must both
            # keep it.
            entry("FB-edge", "FB2", "FBGIU Bakery", "2026-07-31 18:30:00", "Friday, 31 Jul", "6", "102"),
            entry("June", "FB3", "FBGIU Deli", "2026-06-20 10:00:00", "Saturday, 20 Jun", "5", "103"),
        ])
        storage_service.save_cards([{**entry(
            "Pend", "FB4", "FBGIU Bar", "2026-07-10 10:00:00", "Wednesday, 08 Jul", "3", "104",
        ), "status": "Pending"}])

        self.start_str, self.end_str = "2026-07-01", "2026-07-31"
        self.window = date_utils.get_custom_range(
            datetime(2026, 7, 1), datetime(2026, 7, 31))

    def _dashboard_approved_count(self, project_type=None):
        start, end = self.window
        return storage_service.get_status_project_counts(
            start_date=start, end_date=end, project_type=project_type)["approve"]

    def test_export_count_matches_dashboard_approved_card(self):
        count = storage_service.export_act_invoice_overview_range(
            self.start_str, self.end_str, self._export_path())
        self.assertEqual(count, 3)
        self.assertEqual(count, self._dashboard_approved_count())

    def test_export_count_matches_dashboard_per_division(self):
        for project_type in ("beverage", "hospitality"):
            with self.subTest(project_type=project_type):
                count = storage_service.export_act_invoice_overview_range(
                    self.start_str, self.end_str,
                    self._export_path(f"{project_type}.xlsx"), project_type=project_type)
                self.assertEqual(count, self._dashboard_approved_count(project_type))

    def test_exported_rows_match_dashboard_table_rows(self):
        path = self._export_path()
        storage_service.export_act_invoice_overview_range(
            self.start_str, self.end_str, path)

        start, end = self.window
        dashboard_rows = storage_service.get_status_rows(
            "approve", start_date=start, end_date=end)

        ws = load_workbook(path)["Billed Invoice Details"]
        labor = [row for row in ws.iter_rows(min_row=5) if row[9].value == 1]
        self.assertEqual(
            [row[11].value for row in labor],           # Project Number, in file order
            [r["Project Number"] for r in dashboard_rows],  # Dashboard order: Date ASC, subject ASC
        )
        self.assertEqual(
            [row[13].value for row in labor],           # Qty
            [float(r["Qty"]) for r in dashboard_rows],
        )

    def test_manually_entered_rate_flows_into_sales_price(self):
        (record_id,) = self._db_rows(
            "SELECT id FROM timecards_approved WHERE subject = 'FB-A'")[0]
        self.assertTrue(storage_service.update_status_record_field(
            "approve", record_id, "rate", 120.0))

        path = self._export_path()
        storage_service.export_act_invoice_overview_range(
            self.start_str, self.end_str, path, project_type="beverage")
        ws = load_workbook(path)["Billed Invoice Details"]
        prices = {row[11].value: row[14].value
                  for row in ws.iter_rows(min_row=5) if row[9].value == 1}
        self.assertEqual(prices["FB1"], 120.0)
        self.assertEqual(prices["FB2"], 0.0)  # untouched rate exports as 0


class SummaryCsvExportTests(TempDbTestCase):
    """export_summary_csv_range shares the History page's range/division/
    bookkeeping rules but writes a flat CSV -- same filters, and the
    bookkeeping columns (id, is_exported) stay out of the file."""

    def setUp(self):
        super().setUp()
        self._insert_row("timecards_approved", subject="July", project_number="FB1",
                         received="2026-07-05 10:00:00")
        self._insert_row("timecards_approved", subject="June", project_number="FB2",
                         day="Friday, 26 Jun", received="2026-06-26 10:00:00")
        self._insert_row("timecards_pending", subject="Pend", project_number="FB3",
                         received="2026-07-10 10:00:00")

    def _export_and_read(self, start="2026-07-01", end="2026-07-31", project_type=None):
        path = self._export_path("summary.csv")
        count = storage_service.export_summary_csv_range(
            start, end, path, project_type=project_type)
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        return count, rows[0], rows[1:]

    def test_exports_approved_rows_in_range_only(self):
        count, _, data = self._export_and_read()
        self.assertEqual(count, 1)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0][0], "July")  # subject leads the column list

    def test_bookkeeping_columns_stay_out_of_the_file(self):
        _, header, _ = self._export_and_read()
        self.assertNotIn("id", header)
        self.assertNotIn("is_exported", header)
        self.assertIn("Project Number", header)
        self.assertIn("received", header)

    def test_records_history_and_advances_last_export_date(self):
        self._export_and_read()
        self.assertEqual(storage_service.get_export_history()[0][0], "summary.csv")
        self.assertEqual(storage_service.get_last_export_date(), "2026-07-31")


if __name__ == "__main__":
    unittest.main()

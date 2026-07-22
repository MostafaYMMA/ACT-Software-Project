import os
import sqlite3
import sys
import tempfile
import unittest
from unittest.mock import patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "services"))

import storage_service


_PROJECT_NAME_COL = 4  # "Project Name" in the ACT sheet (see _ACT_HEADERS)
_FIRST_DATA_ROW = 5    # title, blank, header, then data


def _project_names_in(path):
    """Every data row's Project Name in an exported ACT sheet -- what the
    division filter is supposed to be selecting on."""
    from openpyxl import load_workbook

    ws = load_workbook(path).active
    names = []
    for row in range(_FIRST_DATA_ROW, ws.max_row + 1):
        value = ws.cell(row=row, column=_PROJECT_NAME_COL).value
        if value:
            names.append(str(value))
    return names


class ActiveExportProjectTypeTests(unittest.TestCase):
    """
    The History page's All / Food & Beverage / Hospitality toggle has to
    reach the rolling export sheet: picking F&B must produce a sheet of
    only FB-prefixed projects, and finalizing it must leave Hospitality's
    own sheet open and untouched.
    """

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.db_path = os.path.join(self.temp_dir.name, "cards_test.db")
        self.exports_dir = os.path.join(self.temp_dir.name, "exports")

        for name, value in (("DB_PATH", self.db_path), ("EXPORTS_DIR", self.exports_dir)):
            patcher = patch.object(storage_service, name, value)
            patcher.start()
            self.addCleanup(patcher.stop)

        storage_service.init_db()
        self._insert_approved("FB Kitchen Fitout", "2026-07-06")
        self._insert_approved("FB Bar Refit", "2026-07-07")
        self._insert_approved("HL Lobby Works", "2026-07-08")

    def _insert_approved(self, project_name, day_date):
        conn = sqlite3.connect(self.db_path)
        conn.execute(
            'INSERT INTO timecards_approved '
            '(day, "Date", subject, sender, "Project Number", "Project Name", '
            '"Task Name", "Qty", rate, period, received) '
            'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            ("Monday", day_date, project_name, "s@x.com", "P1", project_name,
             "T1", "8", "100", "from 2026-07-06 to 2026-07-12", f"{day_date} 10:00:00"),
        )
        conn.commit()
        conn.close()

    def test_food_beverage_sheet_holds_only_fb_rows(self):
        result = storage_service.rebuild_active_export(project_type="beverage")

        self.assertEqual(result["new_rows"], 2)
        self.assertEqual(result["total_rows"], 2)
        self.assertEqual(
            sorted(_project_names_in(result["path"])), ["FB Bar Refit", "FB Kitchen Fitout"]
        )

    def test_hospitality_sheet_holds_only_hl_rows(self):
        result = storage_service.rebuild_active_export(project_type="hospitality")

        self.assertEqual(result["total_rows"], 1)
        self.assertEqual(_project_names_in(result["path"]), ["HL Lobby Works"])

    def test_the_two_divisions_get_separate_files(self):
        beverage = storage_service.rebuild_active_export(project_type="beverage")
        hospitality = storage_service.rebuild_active_export(project_type="hospitality")

        self.assertNotEqual(beverage["path"], hospitality["path"])
        # Rebuilding one must not pull the other's rows into it.
        again = storage_service.rebuild_active_export(project_type="beverage")
        self.assertEqual(again["new_rows"], 0)
        self.assertEqual(again["total_rows"], 2)
        self.assertEqual(
            sorted(_project_names_in(again["path"])), ["FB Bar Refit", "FB Kitchen Fitout"]
        )

    def test_each_division_tracks_its_own_active_path(self):
        beverage = storage_service.rebuild_active_export(project_type="beverage")
        hospitality = storage_service.rebuild_active_export(project_type="hospitality")

        self.assertEqual(storage_service.get_active_export_path("beverage"), beverage["path"])
        self.assertEqual(storage_service.get_active_export_path("hospitality"), hospitality["path"])
        self.assertIsNone(storage_service.get_active_export_path())

    def test_finalizing_one_division_leaves_the_other_open(self):
        storage_service.rebuild_active_export(project_type="beverage")
        hospitality = storage_service.rebuild_active_export(project_type="hospitality")

        storage_service.finalize_active_export("2026-07-31", project_type="beverage")

        # F&B closed: pointer cleared, next Update opens a fresh file.
        self.assertIsNone(storage_service.get_active_export_path("beverage"))
        # Hospitality untouched: same file, same rows, still open.
        self.assertEqual(storage_service.get_active_export_path("hospitality"), hospitality["path"])
        still_open = storage_service.rebuild_active_export(project_type="hospitality")
        self.assertEqual(still_open["path"], hospitality["path"])
        self.assertEqual(still_open["total_rows"], 1)

    def test_a_row_already_in_a_sheet_is_not_picked_up_by_another(self):
        """A timecard is a piece of work invoiced once. Exporting it under
        All must stop the F&B sheet from claiming it as new."""
        storage_service.rebuild_active_export()  # All: takes all three

        beverage = storage_service.rebuild_active_export(project_type="beverage")

        self.assertEqual(beverage["new_rows"], 0)
        self.assertEqual(beverage["total_rows"], 0)


if __name__ == "__main__":
    unittest.main()

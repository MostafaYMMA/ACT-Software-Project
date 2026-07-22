import json
import os
import sqlite3
import sys
import tempfile
import unittest
from unittest.mock import patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "services"))

import storage_service
import sync_payload_excel


# Same shape extract() produces -- see test_current_sheet_color_sync.
_ENTRY = {
    "status": "Approved", "day": "Monday",
    "project_code": "P1", "project_name": "FB Kitchen", "task": "T1",
    "hours": "8", "person_number": "PN1", "subject": "S", "sender": "s@x.com",
    "received": "2026-07-06 10:00:00", "period": "from 2026-07-06 to 2026-07-12",
    "name": "N", "labor_type": "L", "time_type": "T",
}


class _DbTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.db_path = os.path.join(self.temp_dir.name, "cards_test.db")

        patcher = patch.object(storage_service, "DB_PATH", self.db_path)
        patcher.start()
        self.addCleanup(patcher.stop)

        storage_service.init_db()

    def _seed(self, entry=None):
        storage_service.save_cards([entry or dict(_ENTRY)])
        return storage_service.get_current_sheet_rows()[0]["id"]


class ChangeTrackingTests(_DbTestCase):
    """updated_at is the whole basis of the change feed -- if an edit
    doesn't stamp it, that edit silently never reaches the other user."""

    def test_freshly_scanned_row_is_not_a_change(self):
        self._seed()
        self.assertEqual(storage_service.get_current_sheet_changes_since(None), [])

    def test_cell_edit_stamps_and_is_picked_up(self):
        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")

        changes = storage_service.get_current_sheet_changes_since(None)
        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["Qty"], "9")
        self.assertIsNotNone(changes[0]["updated_at"])

    def test_colour_edit_stamps_and_is_picked_up(self):
        row_id = self._seed()
        storage_service.set_current_sheet_row_color(row_id, "#FFD966")

        changes = storage_service.get_current_sheet_changes_since(None)
        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["row_color"], "#FFD966")

    def test_changes_since_excludes_already_sent_edits(self):
        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")
        cutoff = storage_service.get_current_sheet_changes_since(None)[0]["updated_at"]

        self.assertEqual(storage_service.get_current_sheet_changes_since(cutoff), [])

    def test_rows_come_back_whole_not_as_a_diff(self):
        """A row can be brand new to the receiver, so a field-level diff
        would be useless -- every column has to be there."""
        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")

        change = storage_service.get_current_sheet_changes_since(None)[0]
        for column in ("timecard_id", "day", "Date", "Project Number",
                       "Project Name", "Task Name", "name", "person_number"):
            self.assertIn(column, change)
        self.assertEqual(change["Project Name"], "FB Kitchen")


class SyncStateTests(_DbTestCase):
    def test_state_starts_unsynced_at_seq_one(self):
        state = storage_service.get_sync_state()
        self.assertIsNone(state["last_synced_at"])
        self.assertEqual(state["next_seq"], 1)

    def test_advance_moves_marker_and_burns_a_seq(self):
        storage_service.advance_sync_state("2026-07-20 10:00:00")
        state = storage_service.get_sync_state()
        self.assertEqual(state["last_synced_at"], "2026-07-20 10:00:00")
        self.assertEqual(state["next_seq"], 2)

        storage_service.advance_sync_state("2026-07-21 10:00:00")
        self.assertEqual(storage_service.get_sync_state()["next_seq"], 3)

    def test_payload_is_none_when_nothing_changed(self):
        self._seed()
        self.assertIsNone(storage_service.build_current_sheet_sync_payload())

    def test_payload_carries_metadata_and_rows(self):
        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")

        payload = storage_service.build_current_sheet_sync_payload()
        self.assertEqual(payload["seq"], 1)
        self.assertEqual(payload["device_id"], storage_service.get_device_id())
        self.assertTrue(payload["generated_at"])
        self.assertEqual(len(payload["rows"]), 1)

    def test_not_advancing_after_a_failed_send_keeps_the_change_queued(self):
        """The reason advance_sync_state is called only on success: the
        rows must still be in the NEXT payload."""
        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")

        first = storage_service.build_current_sheet_sync_payload()
        # ...send fails, so no advance_sync_state call...
        second = storage_service.build_current_sheet_sync_payload()

        self.assertEqual(len(second["rows"]), 1)
        self.assertEqual(second["seq"], first["seq"])

    def test_advancing_after_a_successful_send_clears_the_queue(self):
        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")

        payload = storage_service.build_current_sheet_sync_payload()
        storage_service.advance_sync_state(payload["cutoff"])

        self.assertIsNone(storage_service.build_current_sheet_sync_payload())


class ApplySyncPayloadTests(_DbTestCase):
    def _payload(self, rows, seq=1, device_id="otherdev1234"):
        return {"kind": "sheet", "device_id": device_id, "seq": seq,
                "generated_at": "2026-07-20 10:00:00", "cutoff": "2026-07-20 10:00:00",
                "rows": rows}

    def _incoming_row(self, **overrides):
        row = {
            "id": 999, "timecard_id": 4242, "day": "Tuesday", "Date": "2026-07-07",
            "Project Number": "P9", "Project Name": "HL Lobby", "Task Name": "T9",
            "Qty": "6", "rate": 50.0, "name": "Remote User", "person_number": "PN9",
            "period": "from 2026-07-06 to 2026-07-12", "subject": "S", "sender": "r@x.com",
            "received": "2026-07-07 09:00:00", "row_color": "#FF0000",
            "color_updated_at": "2026-07-20 09:00:00", "color_updated_by": "otherdev1234",
            "updated_at": "2026-07-20 09:00:00",
        }
        row.update(overrides)
        return row

    def _sheet_rows(self):
        conn = sqlite3.connect(self.db_path)
        try:
            cursor = conn.execute("SELECT * FROM current_sheet")
            columns = [d[0] for d in cursor.description]
            return [dict(zip(columns, r)) for r in cursor.fetchall()]
        finally:
            conn.close()

    def test_unseen_row_is_inserted_whole(self):
        result = storage_service.apply_sync_payload(self._payload([self._incoming_row()]))
        self.assertTrue(result["applied"])
        self.assertEqual(result["inserted"], 1)

        rows = self._sheet_rows()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Project Name"], "HL Lobby")
        self.assertEqual(rows[0]["row_color"], "#FF0000")

    def test_senders_row_id_does_not_overwrite_the_local_key(self):
        """id is a local autoincrement and means nothing on the sender's
        machine -- matching is on timecard_id."""
        self._seed()
        local = self._sheet_rows()[0]
        storage_service.apply_sync_payload(self._payload([self._incoming_row(id=local["id"])]))

        rows = self._sheet_rows()
        self.assertEqual(len(rows), 2)
        self.assertEqual({r["timecard_id"] for r in rows}, {local["timecard_id"], 4242})

    def test_known_row_is_overwritten_in_place(self):
        row_id = self._seed()
        local = self._sheet_rows()[0]
        storage_service.update_current_sheet_field(row_id, "Qty", "1")

        storage_service.apply_sync_payload(self._payload(
            [self._incoming_row(timecard_id=local["timecard_id"], Qty="7")]
        ))

        rows = self._sheet_rows()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Qty"], "7")

    def test_same_seq_applied_twice_is_skipped(self):
        """What stops the same email being applied again if it's rescanned."""
        payload = self._payload([self._incoming_row()])
        storage_service.apply_sync_payload(payload)

        second = storage_service.apply_sync_payload(payload)
        self.assertTrue(second["skipped_duplicate"])
        self.assertFalse(second["applied"])
        self.assertEqual(len(self._sheet_rows()), 1)

    def test_same_seq_from_a_different_sender_is_not_confused_for_a_duplicate(self):
        storage_service.apply_sync_payload(self._payload([self._incoming_row()], seq=1))
        result = storage_service.apply_sync_payload(
            self._payload([self._incoming_row(timecard_id=5151)], seq=1, device_id="thirddev5678")
        )
        self.assertTrue(result["applied"])
        self.assertEqual(len(self._sheet_rows()), 2)

    def test_applied_rows_are_not_echoed_back_to_the_sender(self):
        """A received row must not look locally edited, or it would bounce
        back on the next push forever."""
        storage_service.apply_sync_payload(self._payload([self._incoming_row()]))
        self.assertIsNone(storage_service.build_current_sheet_sync_payload())

    def test_unknown_columns_from_a_newer_build_are_dropped_not_fatal(self):
        result = storage_service.apply_sync_payload(
            self._payload([self._incoming_row(some_future_column="x")])
        )
        self.assertTrue(result["applied"])
        self.assertEqual(self._sheet_rows()[0]["timecard_id"], 4242)


class PayloadSerializationTests(_DbTestCase):
    """The JSON is the data; the .xlsx is decoration. Only the JSON is
    ever read back."""

    def test_json_round_trips_the_payload(self):
        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")
        payload = storage_service.build_current_sheet_sync_payload()

        json_path = os.path.join(self.temp_dir.name, "p.json")
        xlsx_path = os.path.join(self.temp_dir.name, "p.xlsx")
        sync_payload_excel.write_sheet_payload_files(payload, json_path, xlsx_path)

        loaded = sync_payload_excel.read_sheet_payload_json(json_path)
        self.assertEqual(loaded["seq"], payload["seq"])
        self.assertEqual(loaded["device_id"], payload["device_id"])
        self.assertEqual(len(loaded["rows"]), 1)
        self.assertEqual(loaded["rows"][0]["Qty"], "9")

    def test_human_xlsx_is_written_from_the_same_rows(self):
        from openpyxl import load_workbook

        row_id = self._seed()
        storage_service.update_current_sheet_field(row_id, "Qty", "9")
        payload = storage_service.build_current_sheet_sync_payload()

        json_path = os.path.join(self.temp_dir.name, "p.json")
        xlsx_path = os.path.join(self.temp_dir.name, "p.xlsx")
        sync_payload_excel.write_sheet_payload_files(payload, json_path, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb.active
        self.assertEqual(ws.max_row, 2)  # header + the one changed row
        self.assertTrue(ws.cell(row=1, column=1).font.bold)
        self.assertIn("FB Kitchen", [c.value for c in ws[2]])
        wb.close()

    def test_malformed_json_is_skipped_rather_than_half_applied(self):
        bad_path = os.path.join(self.temp_dir.name, "bad.json")
        with open(bad_path, "w", encoding="utf-8") as handle:
            handle.write("{not json")
        self.assertIsNone(sync_payload_excel.read_sheet_payload_json(bad_path))

        wrong_shape = os.path.join(self.temp_dir.name, "wrong.json")
        with open(wrong_shape, "w", encoding="utf-8") as handle:
            json.dump({"seq": 1}, handle)
        self.assertIsNone(sync_payload_excel.read_sheet_payload_json(wrong_shape))


class SyncMailIsNotATimecardTests(unittest.TestCase):
    """The scanner must reject the app's own mail before it ever reaches
    approval detection -- a sync attachment is full of project/task/hours
    words and would otherwise match the loose attachment rule."""

    class _FakeItem:
        def __init__(self, subject):
            self.Subject = subject
            self.UnRead = True

    def test_sync_subjects_are_recognised(self):
        import filter_service

        for subject in (
            "ACT-SYNC v1 | sheet | a1b2c3d4e5f6 | seq=3",
            "ACT-SYNC v1 | snapshot | a1b2c3d4e5f6 | seq=7",
            "ACT-SYNC v2 | something-new | abc | seq=1",
        ):
            self.assertTrue(filter_service.is_sync_mail(self._FakeItem(subject)), subject)

    def test_real_timecard_subjects_are_untouched(self):
        import filter_service

        for subject in (
            "FW: FYI: Time Card Approved",
            "Approved timecard for week of 2026-07-06",
            "Expense Report Approved - Project X",
        ):
            self.assertFalse(filter_service.is_sync_mail(self._FakeItem(subject)), subject)

    def test_sync_mail_is_dropped_by_both_scanners(self):
        import filter_service

        item = self._FakeItem("ACT-SYNC v1 | sheet | a1b2c3d4e5f6 | seq=3")
        counters = filter_service.Counters()
        self.assertIsNone(filter_service.process_email(item, "unused_temp_dir", counters))
        self.assertIsNone(filter_service.process_email_expense(item, "unused_temp_dir"))
        self.assertEqual(counters.total_emails, 0)


if __name__ == "__main__":
    unittest.main()

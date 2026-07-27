"""
Row colours set on the Current Sheet (current_sheet.row_color) used to
never reach the exported invoice workbook at all: _ACT_ROW_COLUMNS only
selected columns off timecards_approved, which has no colour column of
its own -- the colour lives on a different table entirely, joined to it
only by timecard_id.

Covers the fix end to end through the real rebuild_active_export/
export_act_invoice_overview_range paths (not a synthetic call to
_write_act_invoice_workbook), using the real set_current_sheet_row_color
path to set colours, then opening the produced .xlsx with openpyxl and
inspecting actual cell fills -- see storage_service._row_fill_for_color /
_ACT_ROW_COLUMNS / _ACT_DATA_COLUMNS.
"""

import os
import sys
import tempfile
import unittest
from unittest.mock import patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "services"))

import storage_service
from openpyxl import load_workbook


def _entry(day, code, name, task, person, received):
    return {
        "status": "Approved", "day": day, "project_name": name,
        "project_code": code, "task": task, "hours": "8",
        "name": "Jane", "person_number": person,
        "subject": "Time Entries Approved", "sender": "s@x.com",
        "received": received, "labor_type": "L", "time_type": "R",
        "period": "2026-06-29 to 2026-07-05",
        "rate": None, "rate_updated_at": None, "rate_updated_by": None,
    }


class RowColorExportTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.db_path = os.path.join(self.temp_dir.name, "test.db")
        self.exports_dir = os.path.join(self.temp_dir.name, "exports")
        for name, value in (("DB_PATH", self.db_path), ("EXPORTS_DIR", self.exports_dir)):
            p = patch.object(storage_service, name, value)
            p.start()
            self.addCleanup(p.stop)
        fx = patch.object(storage_service, "_cached_usd_rates", return_value={"AED": 3.67})
        fx.start()
        self.addCleanup(fx.stop)
        storage_service.init_db()

        storage_service.save_cards([
            _entry("2026-07-01", "FB-1", "FB Kitchen", "T1", "P1", "2026-07-01 09:00:00"),
            _entry("2026-07-02", "FB-2", "FB Bar", "T2", "P2", "2026-07-02 09:00:00"),
            _entry("2026-07-03", "FB-3", "FB Lounge", "T3", "P3", "2026-07-03 09:00:00"),
        ])
        self.rows_by_name = {r["Project Name"]: r for r in storage_service.get_current_sheet_rows()}

    def _pairs(self, ws, header_row=4):
        pairs = []
        r = header_row + 1
        while ws.cell(row=r, column=11).value == "LABOR":
            pairs.append((r, r + 1, ws.cell(row=r, column=4).value))
            r += 2
        return {name: (labor_r, exp_r) for labor_r, exp_r, name in pairs}

    def _fill_rgb(self, ws, r, c):
        fill = ws.cell(row=r, column=c).fill
        return None if fill is None or fill.fgColor is None else fill.fgColor.rgb

    def test_colored_rows_appear_in_the_rolling_active_export(self):
        set_current_sheet_row_color = storage_service.set_current_sheet_row_color
        self.assertTrue(set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#F44336"))

        result = storage_service.rebuild_active_export()
        wb = load_workbook(result["path"])
        ws = wb.active
        pairs = self._pairs(ws)
        labor_r, exp_r = pairs["FB Kitchen"]

        labor_fill = self._fill_rgb(ws, labor_r, 2)
        exp_fill = self._fill_rgb(ws, exp_r, 2)
        self.assertNotIn(labor_fill, (None, "00000000"))
        self.assertNotIn(exp_fill, (None, "00000000"))
        self.assertEqual(labor_fill, exp_fill, "both rows of a LABOR/Expense pair must share the same fill")
        self.assertNotEqual(labor_fill, "FFF44336", "the fill must be tinted, not the full-saturation source color")

    def test_uncolored_rows_get_no_fill_at_all(self):
        # FB Kitchen colored, FB Bar and FB Lounge left uncolored.
        storage_service.set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#F44336")

        result = storage_service.rebuild_active_export()
        wb = load_workbook(result["path"])
        ws = wb.active
        pairs = self._pairs(ws)

        for name in ("FB Bar", "FB Lounge"):
            labor_r, exp_r = pairs[name]
            self.assertIn(self._fill_rgb(ws, labor_r, 2), (None, "00000000"), f"{name} LABOR row should be unfilled")
            self.assertIn(self._fill_rgb(ws, exp_r, 2), (None, "00000000"), f"{name} Expense row should be unfilled")

    def test_fill_spans_the_full_data_column_range(self):
        storage_service.set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#4CAF50")
        result = storage_service.rebuild_active_export()
        wb = load_workbook(result["path"])
        ws = wb.active
        labor_r, exp_r = self._pairs(ws)["FB Kitchen"]

        labor_fill = self._fill_rgb(ws, labor_r, 2)
        for col in storage_service._ACT_DATA_COLUMNS:
            self.assertEqual(self._fill_rgb(ws, labor_r, col), labor_fill)
            self.assertEqual(self._fill_rgb(ws, exp_r, col), labor_fill)

    def test_header_and_totals_are_never_colored(self):
        storage_service.set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#F44336")
        result = storage_service.rebuild_active_export()
        wb = load_workbook(result["path"])
        ws = wb.active
        pairs = self._pairs(ws)

        # header keeps its own pre-existing fill (light blue), unaffected
        self.assertEqual(self._fill_rgb(ws, 4, 2), "00ADD8E6")

        last_row = max(exp_r for _labor_r, exp_r in pairs.values())
        totals_row = last_row + 1
        self.assertIn(self._fill_rgb(ws, totals_row, 12), (None, "00000000"))

    def test_malformed_stored_color_is_skipped_not_raised(self):
        row_id = self.rows_by_name["FB Kitchen"]["id"]
        conn = __import__("sqlite3").connect(self.db_path)
        conn.execute("UPDATE current_sheet SET row_color = ? WHERE id = ?", ("not-a-color", row_id))
        conn.commit()
        conn.close()

        result = storage_service.rebuild_active_export()  # must not raise
        wb = load_workbook(result["path"])
        ws = wb.active
        labor_r, _exp_r = self._pairs(ws)["FB Kitchen"]
        self.assertIn(self._fill_rgb(ws, labor_r, 2), (None, "00000000"))

    def test_eight_digit_argb_stored_color_is_handled(self):
        storage_service.set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#804A90D9")
        result = storage_service.rebuild_active_export()
        wb = load_workbook(result["path"])
        ws = wb.active
        labor_r, _exp_r = self._pairs(ws)["FB Kitchen"]
        self.assertNotIn(self._fill_rgb(ws, labor_r, 2), (None, "00000000"))

    def test_export_range_button_path_also_carries_colors(self):
        """export_act_invoice_overview_range shares _ACT_ROW_COLUMNS with
        rebuild_active_export -- confirms the LEFT JOIN was added there too
        (required just to keep the shared column list from drifting, since
        both queries must produce the same shape of row tuple)."""
        storage_service.set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#F44336")
        output_path = os.path.join(self.temp_dir.name, "range_export.xlsx")
        storage_service.export_act_invoice_overview_range("2026-07-01", "2026-07-31", output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        labor_r, exp_r = self._pairs(ws)["FB Kitchen"]
        self.assertNotIn(self._fill_rgb(ws, labor_r, 2), (None, "00000000"))

    def test_table_banding_still_present_alongside_manual_fills(self):
        """Confirms the known risk was actually checked, not assumed away:
        the sheet still carries its TableStyleMedium23/showRowStripes
        table alongside the manual per-row fills."""
        storage_service.set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#F44336")
        result = storage_service.rebuild_active_export()
        wb = load_workbook(result["path"])
        ws = wb.active
        table = list(ws.tables.values())[0]
        self.assertEqual(table.tableStyleInfo.name, "TableStyleMedium23")
        self.assertTrue(table.tableStyleInfo.showRowStripes)
        # And the manual fill is still there, i.e. explicit cell formatting
        # was actually written (Excel renders direct cell formatting over
        # table-style banding -- documented Excel behavior; visual
        # confirmation in Excel itself isn't possible in this environment).
        labor_r, _exp_r = self._pairs(ws)["FB Kitchen"]
        self.assertNotIn(self._fill_rgb(ws, labor_r, 2), (None, "00000000"))

    def test_project_type_filter_still_works_after_the_join(self):
        """The LEFT JOIN mustn't break the "Project Name" prefix filter
        used to scope a rolling sheet to one division (both tables have a
        "Project Name" column -- an unqualified reference would be
        ambiguous, or silently wrong, after the join)."""
        storage_service.set_current_sheet_row_color(self.rows_by_name["FB Kitchen"]["id"], "#F44336")
        result = storage_service.rebuild_active_export(project_type="beverage")
        self.assertEqual(result["total_rows"], 3)  # all three are FB-prefixed


if __name__ == "__main__":
    unittest.main()

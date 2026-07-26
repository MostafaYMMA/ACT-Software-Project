"""
Two bugs in finalize_active_export (services/storage_service.py):

  1. Colours set on Current Sheet (current_sheet.row_color) never made it
     into the finalized export -- _write_act_invoice_workbook read only
     from timecards_approved, which has no colour column at all.
  2. current_sheet was never cleared when a row got finalized, so
     "finalized" rows kept sitting in the Current Sheet UI forever, both
     tables drifting apart from what the export actually reflects.
"""

import os
import sqlite3
import sys
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "services"))

import storage_service


def _entry(project_number, project_name, day="2026-07-01", received="2026-07-01 09:00:00"):
    return {
        "status": "Approved", "day": day, "project_name": project_name,
        "project_code": project_number, "task": "T1", "hours": "8",
        "name": "Jane", "person_number": "P1",
        "subject": "Time Entries Approved", "sender": "s@x.com",
        "received": received, "labor_type": "L", "time_type": "R",
        "period": "2026-06-29 to 2026-07-05",
        "rate": None, "rate_updated_at": None, "rate_updated_by": None,
    }


class FinalizeClearsCurrentSheetAndCarriesColorTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.db_path = os.path.join(self.temp_dir.name, "test.db")
        self.exports_dir = os.path.join(self.temp_dir.name, "exports")

        for name, value in (("DB_PATH", self.db_path), ("EXPORTS_DIR", self.exports_dir)):
            patcher = patch.object(storage_service, name, value)
            patcher.start()
            self.addCleanup(patcher.stop)

        rate_patcher = patch.object(storage_service, "_cached_usd_rates", return_value={"AED": 3.67})
        rate_patcher.start()
        self.addCleanup(rate_patcher.stop)

        storage_service.init_db()

    def _current_sheet_row_count(self):
        conn = sqlite3.connect(self.db_path)
        try:
            return conn.execute("SELECT COUNT(*) FROM current_sheet").fetchone()[0]
        finally:
            conn.close()

    def test_finalize_deletes_the_rows_it_closed_out_from_current_sheet(self):
        storage_service.save_cards([_entry("FB-100", "FB Kitchen")])
        self.assertEqual(self._current_sheet_row_count(), 1)

        storage_service.rebuild_active_export()
        storage_service.finalize_active_export("2026-07-31")

        self.assertEqual(self._current_sheet_row_count(), 0)

    def test_a_finalized_row_does_not_come_back_after_a_later_rescan(self):
        """The bug this guards against: finalize deletes the current_sheet
        row, which also removes the timecard_id UNIQUE guard that was
        blocking sync_current_sheet from re-inserting it. Without
        excluding finalized rows explicitly (see sync_current_sheet), the
        very next Scan Inbox -- including one that runs when the app is
        simply reopened -- would silently bring it right back."""
        entry = _entry("FB-100", "FB Kitchen")
        storage_service.save_cards([entry])
        storage_service.rebuild_active_export()
        storage_service.finalize_active_export("2026-07-31")
        self.assertEqual(self._current_sheet_row_count(), 0)

        # Outlook still has the same email; a later scan (app reopen,
        # overlap re-covering the same window, etc) re-imports it.
        storage_service.save_cards([entry])

        self.assertEqual(
            self._current_sheet_row_count(), 0,
            "a finalized row must not reappear in Current Sheet on a later rescan",
        )

    def test_finalize_does_not_touch_current_sheet_rows_from_a_different_division(self):
        """Finalizing Food & Beverage must leave an untouched Hospitality
        row sitting in Current Sheet -- each division closes out
        independently (see finalize_active_export's own docstring)."""
        storage_service.save_cards([
            _entry("FB-100", "FB Kitchen"),
            _entry("HL-200", "HL Hotel"),
        ])
        self.assertEqual(self._current_sheet_row_count(), 2)

        storage_service.rebuild_active_export(project_type="beverage")
        storage_service.finalize_active_export("2026-07-31", project_type="beverage")

        remaining = storage_service.get_current_sheet_rows()
        self.assertEqual(len(remaining), 1)
        self.assertEqual(remaining[0]["Project Name"], "HL Hotel")

    def test_row_color_set_on_current_sheet_carries_into_the_finalized_workbook(self):
        storage_service.save_cards([_entry("FB-100", "FB Kitchen")])
        row_id = storage_service.get_current_sheet_rows()[0]["id"]
        storage_service.set_current_sheet_row_color(row_id, "#FFD966")

        storage_service.rebuild_active_export()
        result = storage_service.finalize_active_export("2026-07-31")

        wb = load_workbook(result["path"])
        ws = wb.active
        # header_row=4 (see _write_act_invoice_workbook) -> first LABOR row is 5.
        fill = ws.cell(row=5, column=4).fill
        self.assertEqual(fill.fgColor.rgb, "00FFD966")

    def test_uncolored_row_gets_no_fill_in_the_finalized_workbook(self):
        storage_service.save_cards([_entry("FB-100", "FB Kitchen")])

        storage_service.rebuild_active_export()
        result = storage_service.finalize_active_export("2026-07-31")

        wb = load_workbook(result["path"])
        ws = wb.active
        fill = ws.cell(row=5, column=4).fill
        self.assertEqual(fill.fill_type, None)


if __name__ == "__main__":
    unittest.main()

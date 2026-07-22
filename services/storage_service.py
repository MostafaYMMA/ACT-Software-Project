import sqlite3
import csv
import json
import os
import re
import urllib.error
import urllib.request
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "data", "cards.db")

_HEADER_FONT = Font(bold=True, size=13, color="000000")
_HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
_BODY_FONT = Font(color="000000")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

_PERIOD_PATTERN = re.compile(
    r"from\s+(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})",
    re.IGNORECASE
)

# --- ACT "Invoice Overview per Period" template constants -------------
_ACT_HEADERS = [
    "Date", "Invoice No", "Project Name", "Period", "Task Name", "Project Mgr",
    "PO", "SOW", "Line", "Type", "Project Number", "Consultant", "Qty",
    "Sales Price", "Total Amount",
]
_ACT_TITLE = "Advanced Computer Technology (Middle East) LABOR and EXPENSE Overview per Period"
_ACT_PO = "AE110007829"
_USD_FORMAT = (
    '_([$USD]\\ * #,##0.00_);_([$USD]\\ * \\(#,##0.00\\);'
    '_([$USD]\\ * "-"??_);_(@_)'
)

# --- Live currency conversion (expenses -> USD) -------------------------
_FX_API_URL = "https://open.er-api.com/v6/latest/USD"  # free, keyless; {code: units-per-USD}
_FX_CACHE_TTL_SECONDS = 3600  # re-fetch at most once an hour
_FX_REQUEST_TIMEOUT_SECONDS = 5


def _ensure_columns(conn, table, column_defs):
    """
    Adds any column in column_defs (list of "name TYPE" strings) that
    doesn't already exist on `table`. Lets older databases created before
    a schema change pick up new columns without losing existing rows.
    """
    existing = {row[1] for row in conn.execute(f'PRAGMA table_info("{table}")')}
    for column_def in column_defs:
        if column_def.startswith('"'):
            column_name = column_def[1:column_def.index('"', 1)]
        else:
            column_name = column_def.split()[0]
        if column_name not in existing:
            conn.execute(f'ALTER TABLE "{table}" ADD COLUMN {column_def}')


STATUS_TABLES = {
    "Approved": "timecards_approved",
    "Pending": "timecards_pending",
    "Rejected": "timecards_rejected",
}

# The business splits its work into two project types, told apart by the code
# the project name starts with: "FB..." (FBGIU ...) is Food & Beverage, "HL..."
# (HLGIU ...) is Hospitality. Each gets its own table (see
# _rebuild_project_type_tables), holding the approved AND pending entries of
# that type -- the two statuses that are still live work. Rejected entries
# belong to neither: they were turned down, so they aren't part of either
# division's book of work. A project named after neither code (e.g. "Global
# Customer Support Platform Upgrade") belongs to no division and lands in
# neither table.
PROJECT_TYPE_TABLES = {
    "beverage": "timecards_food_beverage",
    "hospitality": "timecards_hospitality",
}
PROJECT_TYPE_PREFIXES = {
    "beverage": "FB",
    "hospitality": "HL",
}
PROJECT_TYPE_LABELS = {
    "beverage": "Food & Beverage",
    "hospitality": "Hospitality",
}
# Which status tables feed the project-type tables, and the value each
# contributes to their "status" column.
PROJECT_TYPE_SOURCE_STATUSES = ("Approved", "Pending")


def _project_type_clause(project_type):
    """
    (sql_fragment, params) restricting a query to one project type, by the
    code "Project Name" starts with -- case-insensitively, and tolerant of a
    stray leading space. Returns ("", []) for None or an unknown type, i.e.
    "all project types", so callers can splice it in unconditionally.
    """
    prefix = PROJECT_TYPE_PREFIXES.get(project_type)
    if prefix is None:
        return "", []
    return (
        f'UPPER(SUBSTR(TRIM("Project Name"), 1, {len(prefix)})) = ?',
        [prefix.upper()],
    )


def get_status_project_counts(start_date=None, end_date=None, project_type=None):
    """
    Return row counts for approved, pending, and rejected records. When
    start_date/end_date are given, only rows whose "received" timestamp
    falls within [start_date, end_date] are counted (see date_utils for
    building these from a UI period choice) -- mirrors get_status_rows so
    the Dashboard's stat cards and table always agree on the same window.

    project_type ("beverage"/"hospitality", or None for all) narrows the
    count to one division, by the same project-name rule the project-type
    tables are built on.
    """
    where, params = _project_type_clause(project_type)
    where_sql = f" WHERE {where}" if where else ""

    conn = sqlite3.connect(DB_PATH)
    try:
        counts = {}
        for label, table_name in STATUS_TABLES.items():
            if conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (table_name,),
            ).fetchone() is None:
                counts[label.lower()] = 0
                continue

            if start_date is None and end_date is None:
                row = conn.execute(
                    f'SELECT COUNT(*) FROM "{table_name}"{where_sql}', params
                ).fetchone()
                counts[label.lower()] = row[0] or 0
            else:
                received_values = conn.execute(
                    f'SELECT received FROM "{table_name}"{where_sql}', params
                ).fetchall()
                counts[label.lower()] = sum(
                    1 for (received,) in received_values if _received_in_range(received, start_date, end_date)
                )

        return {
            "approve": counts.get("approved", 0),
            "pending": counts.get("pending", 0),
            "reject": counts.get("rejected", 0),
        }
    finally:
        conn.close()


_STATUS_LABELS = {"approve": "Approved", "pending": "Pending", "reject": "Rejected"}


def get_status_columns(status_key="approve"):
    """
    Returns the column names of a status table, in schema order, without
    reading any rows. Lets the Dashboard show the full set of headings on
    an empty table (before the first scan), where get_status_rows has no
    row to take keys from.
    """
    table_name = STATUS_TABLES.get(_STATUS_LABELS.get(status_key, "Approved"))
    if table_name is None:
        return []

    conn = sqlite3.connect(DB_PATH)
    try:
        return [row[1] for row in conn.execute(f'PRAGMA table_info("{table_name}")')]
    finally:
        conn.close()


def get_status_rows(status_key, start_date=None, end_date=None, project_type=None):
    """
    Return rows for the selected status table as dictionaries. When
    start_date/end_date are given, only rows whose "received" timestamp
    falls within [start_date, end_date] are returned.

    project_type ("beverage"/"hospitality", or None for all) narrows the
    rows to one division. For the Approved and Pending statuses that gives
    exactly what sits in that division's project-type table; Rejected has
    no such table (rejected work belongs to neither division's book), but
    the same project-name rule still applies here so the Dashboard can
    filter all three statuses consistently.
    """
    table_name = STATUS_TABLES.get(_STATUS_LABELS.get(status_key, "Approved"))
    if table_name is None:
        return []

    where, params = _project_type_clause(project_type)
    where_sql = f" WHERE {where}" if where else ""

    conn = sqlite3.connect(DB_PATH)
    try:
        cursor = conn.execute(
            f'SELECT * FROM "{table_name}"{where_sql} ORDER BY "Date" ASC, subject ASC',
            params,
        )
        columns = [desc[0] for desc in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        if start_date is not None or end_date is not None:
            rows = [row for row in rows if _received_in_range(row.get("received"), start_date, end_date)]
        return rows
    finally:
        conn.close()


def get_project_type_rows(project_type, start_date=None, end_date=None):
    """
    Rows straight out of one project-type table (timecards_food_beverage /
    timecards_hospitality) -- approved and pending entries of that division
    together, each carrying a "status" column saying which it is. Unlike
    get_status_rows this reads the derived table itself, so it's the view
    to use when the division, not the status, is what's being reported on.
    """
    table_name = PROJECT_TYPE_TABLES.get(project_type)
    if table_name is None:
        return []

    conn = sqlite3.connect(DB_PATH)
    try:
        if conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,)
        ).fetchone() is None:
            return []

        cursor = conn.execute(
            f'SELECT * FROM "{table_name}" ORDER BY "Date" ASC, subject ASC'
        )
        columns = [desc[0] for desc in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        if start_date is not None or end_date is not None:
            rows = [row for row in rows if _received_in_range(row.get("received"), start_date, end_date)]
        return rows
    finally:
        conn.close()


# Columns search_records is allowed to match against. Also the vocabulary the
# Records page offers in its field picker -- keeping the list here means the
# UI and the WHERE clause can't drift apart (and, since field names get
# interpolated into SQL, acts as the whitelist that keeps arbitrary strings
# out of the query).
SEARCHABLE_FIELDS = [
    "subject", "sender", "Project Number", "Project Name", "Task Name",
    "name", "person_number",
]


def search_records(query="", fields=None):
    """
    Searches across all three status tables and returns full rows (every
    column, SELECT *). An empty query returns everything. Each result dict
    is tagged with "status" (Approved/Pending/Rejected) so callers can
    show where a match currently stands.

    fields limits which columns the query is matched against; it must be
    a subset of SEARCHABLE_FIELDS (anything else is ignored). None or an
    empty list means all of them -- the pre-field-picker behavior.
    """
    if fields:
        fields = [f for f in fields if f in SEARCHABLE_FIELDS]
    if not fields:
        fields = SEARCHABLE_FIELDS

    conn = sqlite3.connect(DB_PATH)
    try:
        results = []
        term = query.strip()
        like = f"%{term}%"
        for status_label, table_name in STATUS_TABLES.items():
            if conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,)
            ).fetchone() is None:
                continue

            if term:
                where = " OR ".join(f'"{field}" LIKE ?' for field in fields)
                cursor = conn.execute(
                    f'SELECT * FROM "{table_name}" WHERE {where}',
                    [like] * len(fields),
                )
            else:
                cursor = conn.execute(f'SELECT * FROM "{table_name}"')

            columns = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                record = dict(zip(columns, row))
                record["status"] = status_label
                results.append(record)

        results.sort(key=lambda r: r.get("Date") or "", reverse=True)
        return results
    finally:
        conn.close()


def update_status_record_field(status_key, record_id, column, value):
    """
    Writes one cell edit from the Dashboard grid back to the record's row
    (identified by id) in the status table status_key maps to. The column
    must actually exist on the table -- column names can't be bound as SQL
    parameters, so this check is what keeps the interpolation safe -- and
    "id" itself is refused. Returns True when a row was updated; False for
    an unknown column/table, a missing id, or an edit the table's UNIQUE
    constraint rejects (e.g. making a row a duplicate of another).
    """
    table_name = STATUS_TABLES.get(_STATUS_LABELS.get(status_key, "Approved"))
    if table_name is None or column in ("id", "received_month"):
        return False

    conn = sqlite3.connect(DB_PATH)
    try:
        valid_columns = {row[1] for row in conn.execute(f'PRAGMA table_info("{table_name}")')}
        if column not in valid_columns:
            return False
        try:
            if column == "received":
                # received_month is derived from received and is half of what
                # makes a row unique -- editing one without the other would
                # leave the row keyed under a month it isn't in.
                cursor = conn.execute(
                    f'UPDATE "{table_name}" SET received = ?, received_month = ? WHERE id = ?',
                    (value, _received_month(value), record_id),
                )
            elif column == "rate":
                # rate is the one field either user might edit on ANY
                # record, including one the other device originally
                # scanned -- stamping who/when here keeps an audit trail
                # of the last edit, and rides along into the SharePoint
                # current-sheet xlsx (see SNAPSHOT_COLUMNS below).
                cursor = conn.execute(
                    f'UPDATE "{table_name}" SET rate = ?, rate_updated_at = ?, rate_updated_by = ? WHERE id = ?',
                    (value, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), get_device_id(), record_id),
                )
            else:
                cursor = conn.execute(
                    f'UPDATE "{table_name}" SET "{column}" = ? WHERE id = ?',
                    (value, record_id),
                )
        except sqlite3.IntegrityError:
            return False
        if cursor.rowcount > 0:
            # The division tables are copies of these rows -- a cell edit here
            # (a rate typed in, or a "Project Name" corrected, which can move
            # an entry from one division to the other or out of both) has to be
            # carried into them now. Waiting for the next scan would leave them
            # showing the pre-edit value in the meantime.
            _rebuild_project_type_tables(conn)
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()


def find_record_id(status_label, day, project_number, task_name, person_number, received_month):
    """
    Looks up a status table row's id by its natural key -- (day,
    "Project Number", "Task Name", person_number, received_month), the
    same tuple _save_row's UNIQUE constraint uses. For a caller (the
    SharePoint View Current window, see ui/Pages/History.py) that only
    has a row from a merged current-sheet xlsx -- which carries no DB id
    at all -- and needs to resolve it back to the local record it
    corresponds to before it can call update_status_record_field. Returns
    None if no such record exists locally (e.g. the row belongs to
    another device and was never scanned by this one).
    """
    table_name = STATUS_TABLES.get(status_label)
    if table_name is None:
        return None

    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            f'SELECT id FROM "{table_name}" WHERE day = ? AND "Project Number" = ? '
            f'AND "Task Name" = ? AND person_number = ? AND received_month = ?',
            (day, project_number, task_name, person_number, received_month),
        ).fetchone()
        return row[0] if row else None
    finally:
        conn.close()


def _parse_received(value):
    """Best-effort parse of the "received" column back into a naive
    datetime, tolerant of the various formats it may have been stored in
    (see _to_row -- it comes straight from Outlook's item.ReceivedTime)."""
    if not value:
        return None
    text = str(value)
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        try:
            parsed = datetime.strptime(text[:19], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None
    return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed


def _received_in_range(received_value, start_date, end_date):
    """True if the stored "received" text falls within [start_date,
    end_date] (either bound may be None for unbounded). A row whose
    "received" can't be parsed is excluded once a range is in effect --
    there's no timestamp to judge it by."""
    parsed = _parse_received(received_value)
    if parsed is None:
        return False
    if start_date is not None and parsed < start_date:
        return False
    if end_date is not None and parsed > end_date:
        return False
    return True


def get_stale_status_counts(min_age_hours):
    """
    Counts Pending/Rejected rows whose "received" timestamp is at least
    min_age_hours in the past. Backs the Settings-configurable "N
    requests pending/rejected for X time" notification banner.
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        now = datetime.now()
        counts = {"pending": 0, "rejected": 0}
        for status_label in ("Pending", "Rejected"):
            table_name = STATUS_TABLES[status_label]
            if conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,)
            ).fetchone() is None:
                continue
            for (received,) in conn.execute(f'SELECT received FROM "{table_name}"'):
                parsed = _parse_received(received)
                if parsed is None:
                    continue
                age_hours = (now - parsed).total_seconds() / 3600
                if age_hours >= min_age_hours:
                    counts[status_label.lower()] += 1
        counts["total"] = counts["pending"] + counts["rejected"]
        return counts
    finally:
        conn.close()


def get_stale_records(min_age_hours):
    """
    Like get_stale_status_counts, but returns the actual Pending/Rejected
    rows (not just counts), each tagged with "status" and "age_hours".
    Backs the Late tab's "pending/rejected for X days/weeks" list.
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        now = datetime.now()
        results = []
        for status_label in ("Pending", "Rejected"):
            table_name = STATUS_TABLES[status_label]
            if conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,)
            ).fetchone() is None:
                continue
            cursor = conn.execute(f"""
                SELECT subject, sender, "Project Number", "Project Name", "Task Name",
                       "Date", "Qty", received
                FROM "{table_name}"
            """)
            columns = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                record = dict(zip(columns, row))
                parsed = _parse_received(record.get("received"))
                if parsed is None:
                    continue
                age_hours = (now - parsed).total_seconds() / 3600
                if age_hours >= min_age_hours:
                    record["status"] = status_label
                    record["age_hours"] = age_hours
                    results.append(record)

        results.sort(key=lambda r: r["age_hours"], reverse=True)
        return results
    finally:
        conn.close()


def _create_timecards_table(conn, table_name):
    # What makes two entries the SAME entry: the same work day, on the same
    # project and task, from the same person, received in the same month.
    # received_month is in the key on purpose -- the same timecard received
    # in a different month is a separate entry, not a duplicate to overwrite.
    # It's a stored column rather than an expression because SQLite can't put
    # substr(received, 1, 7) inside a table-level UNIQUE constraint.
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS "{table_name}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            day TEXT,
            "Date" TEXT,
            labor_type TEXT,
            time_type TEXT,
            "Qty" TEXT,
            "Project Number" TEXT,
            "Project Name" TEXT,
            "Task Name" TEXT,
            name TEXT,
            period TEXT,
            person_number TEXT,
            subject TEXT,
            sender TEXT,
            received TEXT,
            received_month TEXT,
            is_exported INTEGER DEFAULT 0,
            rate REAL DEFAULT 0,
            ----UNIQUE(day, "Project Number", "Task Name", sender, received_month)
            ----UNIQUE(day, "Project Number","Project Name",person_number ,"Task Name", received_month)
            UNIQUE(day, "Project Number", "Task Name", person_number, received_month)

        )
    """)


def _needs_received_month_rebuild(conn, table_name):
    """
    True when table_name exists but predates received_month. Such a table was
    created before the column joined the row key, and can't be patched in place
    (see _rebuild_status_table_with_received_month). A table that doesn't exist
    yet returns False -- _create_timecards_table builds it correctly from the
    start.
    """
    columns = {row[1] for row in conn.execute(f'PRAGMA table_info("{table_name}")')}
    return bool(columns) and "received_month" not in columns


def _rebuild_status_table_with_received_month(conn, table_name):
    """
    Brings a status table created before received_month existed up to the
    current schema. received_month is half of the row key and part of the
    table's UNIQUE constraint, so ALTER TABLE ADD COLUMN can't introduce it:
    that can't extend a UNIQUE constraint, and _upsert_timecard's ON CONFLICT
    needs the constraint to match. The table is rebuilt instead -- renamed
    aside, a fresh one created, existing rows copied over with received_month
    derived from each row's received timestamp, then the old table dropped.

    Row ids are carried across so invoice_lines.timecard_id links back into the
    approved table still resolve. When two old rows collapse onto the new
    (day, project, task, person, month) key, the one with the latest received
    wins -- the same "newest state is current" rule _save_row enforces.
    """
    tmp_name = f"{table_name}__pre_received_month"
    conn.execute(f'DROP TABLE IF EXISTS "{tmp_name}"')
    conn.execute(f'ALTER TABLE "{table_name}" RENAME TO "{tmp_name}"')
    _create_timecards_table(conn, table_name)

    old_columns = {row[1] for row in conn.execute(f'PRAGMA table_info("{tmp_name}")')}
    new_columns = [row[1] for row in conn.execute(f'PRAGMA table_info("{table_name}")')]

    # Copy every column the new table has that the old one still carries, apart
    # from received_month, which is derived below rather than copied.
    carried = [c for c in new_columns if c in old_columns and c != "received_month"]
    select_exprs = []
    for column in carried:
        # person_number is part of the key and must never be NULL (see _to_row);
        # an old row that predates the column needs a concrete '' instead.
        if column == "person_number":
            select_exprs.append("COALESCE(\"person_number\", '')")
        else:
            select_exprs.append(f'"{column}"')

    has_received = "received" in old_columns
    received_select = 'substr("received", 1, 7)' if has_received else "NULL"
    order_sql = 'ORDER BY "received"' if has_received else ""
    carried_sql = ", ".join(f'"{c}"' for c in carried)
    select_sql = ", ".join(select_exprs)

    conn.execute(f"""
        INSERT OR REPLACE INTO "{table_name}" ({carried_sql}, received_month)
        SELECT {select_sql}, {received_select}
        FROM "{tmp_name}"
        {order_sql}
    """)
    conn.execute(f'DROP TABLE IF EXISTS "{tmp_name}"')


def _create_project_type_table(conn, table_name):
    # A division's book of live work: the approved and pending entries whose
    # project name starts with that division's letter, copied out of the two
    # status tables. status says which of the two an entry came from, and
    # source_id is its id in that table (the pair is unique -- an entry exists
    # in exactly one status table at a time, see _save_row).
    #
    # It's a copy, not a view, so a division can be queried, exported and
    # reported on on its own. Nothing writes to it directly: it is emptied and
    # rebuilt from the status tables on every save (_rebuild_project_type_tables),
    # which is what keeps it from drifting out of step with them.
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS "{table_name}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_id INTEGER,
            status TEXT,
            day TEXT,
            "Date" TEXT,
            labor_type TEXT,
            time_type TEXT,
            "Qty" TEXT,
            "Project Number" TEXT,
            "Project Name" TEXT,
            "Task Name" TEXT,
            name TEXT,
            period TEXT,
            person_number TEXT,
            subject TEXT,
            sender TEXT,
            received TEXT,
            received_month TEXT,
            is_exported INTEGER DEFAULT 0,
            rate REAL DEFAULT 0,
            UNIQUE(status, source_id)
        )
    """)


# The columns a project-type row carries over from its status table, in the
# order both the INSERT and the SELECT below use.
_PROJECT_TYPE_COPIED_COLUMNS = [
    "day", '"Date"', "labor_type", "time_type", '"Qty"', '"Project Number"',
    '"Project Name"', '"Task Name"', "name", "period", "person_number",
    "subject", "sender", "received", "received_month", "is_exported", "rate",
]


def _rebuild_project_type_tables(conn):
    """
    Recomputes both project-type tables from scratch out of the Approved and
    Pending status tables, routing each entry by the first letter of its
    project name (B -> Food & Beverage, H -> Hospitality).

    Rebuilt rather than appended to: an entry that changes status, gets its
    project name corrected, or is dropped from its status table entirely
    would otherwise leave a stale copy behind in the division tables. An
    entry whose project name starts with neither letter belongs to no
    division and simply appears in neither table.
    """
    for project_type, table_name in PROJECT_TYPE_TABLES.items():
        # Same rule the Dashboard's project-type filter runs on, so the tables
        # and the filtered views of the status tables can't disagree.
        where, where_params = _project_type_clause(project_type)
        conn.execute(f'DELETE FROM "{table_name}"')
        for status_label in PROJECT_TYPE_SOURCE_STATUSES:
            source_table = STATUS_TABLES[status_label]
            # Only copy the columns the source actually has. A database created
            # before a column was added to the status tables (received_month is
            # the live case -- _ensure_columns never backfilled it) still has
            # rows worth routing into a division; the missing column is simply
            # left NULL here rather than failing the whole rebuild.
            source_columns = {row[1] for row in conn.execute(f'PRAGMA table_info("{source_table}")')}
            copied = [
                column for column in _PROJECT_TYPE_COPIED_COLUMNS
                if column.strip('"') in source_columns
            ]
            if not copied:
                continue
            columns_sql = ", ".join(copied)
            conn.execute(f"""
                INSERT INTO "{table_name}" (source_id, status, {columns_sql})
                SELECT id, ?, {columns_sql}
                FROM "{source_table}"
                WHERE {where}
            """, [status_label, *where_params])


def _create_expenses_table(conn):
    # Expense reports (see extractor_service.extract_expense): one row per
    # report, not per line item -- the report's own header fields are all we
    # keep, since the Expense Item line table is no longer parsed. "Amount"/
    # currency hold the report's "Report Total : <amount> <currency>" line.
    #
    # "Project Number"/"Project Name" are named exactly like the timecard
    # tables so the two can be joined on "Project Number". What makes a
    # report unique is its own report number: re-reading the same report
    # updates its row rather than duplicating it.
    #
    # A table from before this line-item -> report-level rework has "item"
    # in its schema and no UNIQUE(expense_report) -- there's no way to keep
    # its rows (they were keyed per line, not per report), so it's dropped
    # and rebuilt fresh rather than migrated.
    old_columns = {row[1] for row in conn.execute('PRAGMA table_info("expenses")')}
    if "item" in old_columns:
        conn.execute('DROP TABLE "expenses"')

    conn.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expense_report TEXT,
            "Project Number" TEXT,
            "Project Name" TEXT,
            "Amount" REAL,
            currency TEXT,
            status TEXT,
            submitted_by TEXT,
            subject TEXT,
            sender TEXT,
            received TEXT,
            received_month TEXT,
            UNIQUE(expense_report)
        )
    """)
    # Which timecard RECORD this expense is billed against (see
    # storage_service._fill_sender_timecard_match): link_method is
    # "period_match" (one of the sender's own timecard records has a
    # period containing this report's expense_date), "latest_fallback"
    # (no record's period contains it -- or expense_date is unknown -- so
    # this is simply that sender's most recently received record), or
    # "none" (the sender submitted no timecard record at all, or the
    # report has no sender). matched_timecard_id/matched_day/matched_date/
    # matched_project_number/matched_project_name/matched_task_name are
    # kept as columns for backward compatibility but are no longer
    # populated: a record can span several days/projects/tasks, so none of
    # those has one single value at record level anymore.
    _ensure_columns(conn, "expenses", [
        "link_method TEXT",
        "matched_timecard_id INTEGER",
        "matched_sender TEXT",
        "matched_person_number TEXT",
        "matched_name TEXT",
        "matched_period TEXT",
        "matched_subject TEXT",
        # Completes the (sender, period, subject, received) key that
        # identifies one timecard RECORD (see _fill_sender_timecard_match) --
        # without it, two records from the same sender/period/subject (e.g.
        # the same week resent) couldn't be told apart when resolving a
        # linked expense back to its day-rows for invoicing.
        "matched_received TEXT",
        "matched_day TEXT",
        "matched_project_number TEXT",
        "matched_project_name TEXT",
        "matched_task_name TEXT",
        # expense_date is the LATEST date among this report's own Expense
        # Item lines (see extractor_service._expense_header_fields) --
        # the anchor _fill_sender_timecard_match checks against each
        # candidate record's period. matched_date is no longer populated
        # (see above) -- kept as a column for backward compatibility.
        "expense_date TEXT",
        # Earliest of the report's own Expense Item dates (see
        # extractor_service._expense_header_fields) -- together with
        # expense_date (the latest) these bound the report's own span, which
        # _fill_sender_timecard_match requires a candidate record's period to
        # fully contain.
        "expense_date_start TEXT",
        "matched_date TEXT",
        # "Amount"/currency are the report's own reported total, kept as-is
        # for traceability; amount_usd is that same total converted with
        # the live rate in effect when the report was scanned (see
        # _to_usd) -- what every cross-currency sum (e.g. two reports
        # linked to the same record) actually adds together.
        "amount_usd REAL",
    ])


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)

    # timecards used to hold approved entries only (that was the only
    # status ever fetched). It's renamed aside into timecards_approved so
    # existing data isn't lost, now that pending/rejected get their own
    # tables too.
    existing_tables = {
        row[0] for row in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )
    }
    if "timecards" in existing_tables and "timecards_approved" not in existing_tables:
        conn.execute("ALTER TABLE timecards RENAME TO timecards_approved")

    for table_name in STATUS_TABLES.values():
        # A table created before received_month existed can't be patched in
        # place (the column is part of the UNIQUE key) -- rebuild it first so
        # the CREATE/ensure below act on a table that already has the column.
        if _needs_received_month_rebuild(conn, table_name):
            _rebuild_status_table_with_received_month(conn, table_name)
        _create_timecards_table(conn, table_name)
        # ADD COLUMN with a DEFAULT backfills existing rows with it, so
        # records created before these columns existed start as
        # not-exported / rate 0 rather than NULL. Neither column is in
        # _upsert_timecard's ON CONFLICT update list, so a re-scan never
        # resets an is_exported flag or a manually entered rate.
        _ensure_columns(conn, table_name, [
            "name TEXT", "period TEXT", "person_number TEXT",
            "is_exported INTEGER DEFAULT 0", "rate REAL DEFAULT 0",
            # origin: which device's own Outlook scan this row came from
            # (see get_device_id/save_cards) -- None/legacy rows are
            # treated as this device's own the first time anything
            # touches them (see _save_row's backfill). What this actually
            # gates: build_outgoing_snapshot only ever publishes rows
            # whose origin is THIS device, so a synced-in row is never
            # re-broadcast back out -- without that, two devices would
            # keep re-mailing each other's data back and forth forever.
            "origin TEXT",
            # rate is hand-typed and never touched by a rescan (see
            # _upsert_timecard) -- these two are an audit trail of the
            # last edit (who/when), and ride along into the SharePoint
            # current-sheet xlsx (see SNAPSHOT_COLUMNS below).
            "rate_updated_at TEXT", "rate_updated_by TEXT",
            # highlight_color: a purely local, per-device visual tag (a
            # hex color or empty/NULL for none) set from the SharePoint
            # View Current window -- see ui/Pages/History.py's
            # _CurrentSheetDialog. Deliberately NOT part of
            # SNAPSHOT_COLUMNS / the shared current-sheet xlsx -- it's
            # this device's own view preference, not data to sync.
            "highlight_color TEXT",
        ])

    for table_name in PROJECT_TYPE_TABLES.values():
        _create_project_type_table(conn, table_name)

    _create_expenses_table(conn)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS timecards_summary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            day TEXT,
            "Period" TEXT,
            "Date" TEXT,
            labor_type TEXT,
            time_type TEXT,
            "Qty" REAL,
            "Project Number" TEXT,
            "Project Name" TEXT,
            "Task Name" TEXT,
            "Name" TEXT,
            "Person Number" TEXT,
            subject TEXT,
            sender TEXT,
            received TEXT
        )
    """)
    _ensure_columns(conn, "timecards_summary", ['"Name" TEXT', '"Person Number" TEXT'])

    # invoice_lines used to be one row per (Project Number, Task Name,
    # Period). It's now one row per raw timecard entry (day-level),
    # identified internally via timecard_id -- an older-shape table (no
    # timecard_id column) is renamed aside rather than dropped, so any
    # manually-entered data isn't lost, then rebuilt fresh below.
    existing_invoice_columns = {row[1] for row in conn.execute('PRAGMA table_info("invoice_lines")')}
    if existing_invoice_columns and "timecard_id" not in existing_invoice_columns:
        conn.execute("ALTER TABLE invoice_lines RENAME TO invoice_lines_period_level_backup")

    # Invoicing worksheet: one row per raw timecard entry (approved only).
    # Only the columns we can derive automatically are ever populated by
    # sync code (see _sync_invoice_lines below) -- everything else here is
    # filled in manually later and is never touched by a re-sync.
    # timecard_id links back to timecards_approved.id purely so
    # re-syncing can tell which entries already have a row here, without
    # touching any manual edits.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS invoice_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timecard_id INTEGER UNIQUE,
            "Date" TEXT,
            "Invoice Number" TEXT,
            "Project Name" TEXT,
            "Period" TEXT,
            "Task Name" TEXT,
            "Project Mgr" TEXT,
            "PO" TEXT,
            "SOW" TEXT,
            "Line" TEXT,
            "Type" TEXT,
            "Project Number" TEXT,
            "Consultant" TEXT,
            "Qty" REAL,
            "Sales Price" REAL,
            "Total Amount" REAL GENERATED ALWAYS AS ("Qty" * "Sales Price")
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS export_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            date TEXT
        )
    """)

    # Which approved timecard rows have been written into a rolling export
    # sheet, and which sheet (see rebuild_active_export). finalized = 0 is
    # the file being filled right now; Finalize flips its rows to 1 rather
    # than deleting them, because a row must never come back as "new" for
    # the NEXT sheet just because the sheet it went into got closed.
    #
    # Deliberately not reusing is_exported for that: is_exported is also
    # set by the ad-hoc "Export Last Month"/"Export Range" buttons, so
    # keying off it would let a one-off export silently withhold rows from
    # the rolling sheet.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS active_export_rows (
            timecard_id INTEGER PRIMARY KEY,
            added_at TEXT
        )
    """)
    _ensure_columns(conn, "active_export_rows", [
        "export_path TEXT",
        "finalized INTEGER DEFAULT 0",
    ])

    # A single-row-per-key store for small pieces of app state that must
    # survive between runs. Currently holds one key, 'last_export_date':
    # the received-"to" date of the most recent range export, overwritten on
    # every export so the History page can offer a "From: last export" range.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS app_state (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)

    # Backfill: the project-type tables are otherwise only written on a save,
    # so on a database that already holds approved/pending records from before
    # they existed they'd sit empty until the next scan. Rebuilding here fills
    # them from what's already stored; on an up-to-date database it recomputes
    # the same rows, which is harmless.
    _rebuild_project_type_tables(conn)

    conn.commit()
    conn.close()


# Ensure the database exists and all expected tables are present as soon as
# the storage layer is imported.
init_db()


def _record_export(conn, name: str):
    conn.execute(
        "INSERT INTO export_history (name, date) VALUES (?, ?)",
        (name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )


def get_export_history():
    """Returns (name, date) rows for every export ever done, newest first."""
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute(
        "SELECT name, date FROM export_history ORDER BY date DESC, id DESC"
    ).fetchall()
    conn.close()
    return rows


def _set_last_export_date(conn, end_date: str):
    """
    Overwrites the stored "last export received-to date" with end_date
    (a 'YYYY-MM-DD' string). Single row, keyed 'last_export_date', so each
    export replaces the previous value rather than accumulating history --
    the export_history table already keeps the full log. Takes an open conn
    so it can ride inside the export's own transaction/commit.
    """
    conn.execute(
        "INSERT INTO app_state (key, value) VALUES ('last_export_date', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (end_date,),
    )


def get_device_id():
    """
    A short random id identifying THIS installation, generated once and
    then persisted in app_state forever after. This is what tags a row's
    "origin" (see init_db's _ensure_columns call) and names this device's
    current_<device_id>.xlsx in the SharePoint folder (see
    services/sharepoint_service.py), so each device only ever writes its
    own file regardless of how many other devices share the folder.
    """
    import uuid

    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            "SELECT value FROM app_state WHERE key = 'device_id'"
        ).fetchone()
        if row and row[0]:
            return row[0]
        new_id = uuid.uuid4().hex[:12]
        conn.execute(
            "INSERT INTO app_state (key, value) VALUES ('device_id', ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (new_id,),
        )
        conn.commit()
        return new_id
    finally:
        conn.close()


def _next_outgoing_seq(conn):
    """
    Increments and returns THIS device's outgoing snapshot counter.
    Every call to build_outgoing_snapshot() burns one of these -- it's
    what lets the other side tell "I already have everything through
    seq 8" apart from "here's a stale seq 5 that arrived late", so a pile
    of queued sync mails collapses to "apply only the newest one" with no
    reordering logic needed on the receiving end (see
    apply_incoming_snapshot).
    """
    row = conn.execute(
        "SELECT value FROM app_state WHERE key = 'sync_outgoing_seq'"
    ).fetchone()
    current = int(row[0]) if row and row[0] else 0
    new_value = current + 1
    conn.execute(
        "INSERT INTO app_state (key, value) VALUES ('sync_outgoing_seq', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (str(new_value),),
    )
    return new_value


def get_last_applied_seq(sender_device_id):
    """The highest snapshot seq already applied FROM sender_device_id, or
    0 if none has ever been applied. Per-sender, so this device can sync
    with more than one other device without them stepping on each other's
    sequence numbers."""
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            "SELECT value FROM app_state WHERE key = ?",
            (f"sync_applied_seq:{sender_device_id}",),
        ).fetchone()
        return int(row[0]) if row and row[0] else 0
    finally:
        conn.close()


def _set_last_applied_seq(conn, sender_device_id, seq):
    conn.execute(
        "INSERT INTO app_state (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (f"sync_applied_seq:{sender_device_id}", str(seq)),
    )


def get_last_export_date():
    """
    The received-"to" date ('YYYY-MM-DD') of the most recent range export,
    or None if nothing has been exported yet. Backs the History page's
    "From: last export" button, which uses it as the range's start.
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            "SELECT value FROM app_state WHERE key = 'last_export_date'"
        ).fetchone()
        return row[0] if row else None
    finally:
        conn.close()


def get_last_scan_time():
    """
    The high-water mark for inbox scanning: the ReceivedTime of the newest
    email that was in the folder the last time a scan completed, as a naive
    datetime (second precision), or None if this device has never completed
    a scan.

    sync_service.sync_cards starts each scan from this instead of walking
    the whole inbox -- see set_last_scan_time for why it's stored at all
    rather than derived from the saved records' "received" column.
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            "SELECT value FROM app_state WHERE key = 'last_scan_received'"
        ).fetchone()
        return _parse_received(row[0]) if row and row[0] else None
    finally:
        conn.close()


def set_last_scan_time(received_at):
    """
    Moves the scan high-water mark to `received_at` (a datetime).

    Deliberately the newest email IN THE FOLDER, not the newest one that
    matched the timecard filter: a week of unrelated mail after the last
    real timecard would otherwise be re-walked on every single scan. The
    cost is that this can only ever be trusted as "everything up to here
    has been looked at", which is why sync_cards re-scans a small overlap
    behind it rather than starting exactly here.

    Only ever moves FORWARD -- a caller handing over an older timestamp
    (an empty folder, a failed read) leaves the stored mark alone, so a
    bad read can't quietly force a full re-scan or, worse, be combined
    with the overlap into a window that skips mail.
    """
    if received_at is None:
        return
    if received_at.tzinfo is not None:
        received_at = received_at.replace(tzinfo=None)

    current = get_last_scan_time()
    if current is not None and received_at <= current:
        return

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute(
            "INSERT INTO app_state (key, value) VALUES ('last_scan_received', ?) "
            "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (received_at.strftime("%Y-%m-%d %H:%M:%S"),),
        )
        conn.commit()
    finally:
        conn.close()


def _received_date_only(received: str) -> str:
    """
    Truncates the raw "received" timestamp (item.ReceivedTime, e.g.
    '2026-07-03 09:14:22') down to just the date part, 'YYYY-MM-DD'.
    Used as the "Date" column instead of the timecard's own work-day -
    the record should be dated by when the email arrived, not by which
    day of the timecard it happens to describe (matches how
    invoice_lines' "Date" already works in _sync_invoice_lines below).
    """
    if not received:
        return None
    return str(received)[:10]


def _parse_period(subject: str) -> str:
    """
    Extracts the timecard's full week range from the subject line, e.g.
    'Your Time Entries from 2026-06-27 to 2026-07-03 Were Approved'
    -> '2026-06-27 to 2026-07-03'
    """
    match = _PERIOD_PATTERN.search(subject or "")
    if match:
        return f"{match.group(1)} to {match.group(2)}"
    return None


def _join_distinct(values) -> str:
    seen = []
    for v in values:
        if v and v not in seen:
            seen.append(v)
    return ", ".join(seen)


def _received_month(received: str) -> str:
    """
    The 'YYYY-MM' the entry was received in, taken off the front of the raw
    "received" timestamp. Half of what makes a row unique -- see
    _create_timecards_table.
    """
    if not received:
        return None
    return str(received)[:7]


def _to_row(entry: dict) -> tuple:
    return (
        entry.get("day"),
        _received_date_only(entry.get("received")),
        entry.get("labor_type"),
        entry.get("time_type"),
        entry.get("hours"),
        entry.get("project_code"),
        entry.get("project_name"),
        entry.get("task"),
        entry.get("name"),
        entry.get("period"),
        # Never NULL: person_number is part of the UNIQUE key, and SQLite
        # counts every NULL as distinct from every other -- one missing
        # person number would quietly exempt that row from the whole
        # duplicate/transition rule (and "person_number = NULL" never
        # matches in _find_existing either). The extractor returns None
        # whenever a timecard uses the labelled "Person Number: 12345"
        # layout instead of the bare header its regex expects, so this is
        # a live case, not a theoretical one.
        entry.get("person_number") or "",
        entry.get("subject"),
        entry.get("sender"),
        entry.get("received"),
        _received_month(entry.get("received")),
    )


def _row_key(row: tuple) -> tuple:
    """
    (day, Project Number, Task Name, person_number, received_month) out of a
    _to_row tuple -- the status tables' UNIQUE key. It deliberately says
    nothing about status, so it identifies the same entry across all three
    of them.

    The entry is keyed by WHOSE timecard it is (person_number), not by who
    happened to send the email -- the same timecard forwarded by two people
    is one entry, not two.
    """
    return (row[0], row[5], row[7], row[10], row[14])


def _rebuild_summary(conn):
    """
    Recomputes timecards_summary from scratch out of the raw per-day
    timecards_approved table (summary/export stays approved-only). Merge
    key: sender + subject + Project Number + Task Name (i.e. same weekly
    timecard email, same project, same task).
    """
    cursor = conn.execute('SELECT * FROM timecards_approved')
    columns = [desc[0] for desc in cursor.description]
    rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    groups = {}
    for row in rows:
        key = (row["sender"], row["subject"], row["Project Number"], row["Task Name"])
        groups.setdefault(key, []).append(row)

    merged_rows = []
    for (sender, subject, project_number, task_name), group in groups.items():
        days = sorted(
            {r["day"] for r in group if r["day"]},
            key=lambda d: _DAY_ORDER.index(d.split(",")[0].strip()) if d.split(",")[0].strip() in _DAY_ORDER else 99
        )
        dates = sorted(r["Date"] for r in group if r["Date"])

        total_qty = 0.0
        for r in group:
            try:
                total_qty += float(r["Qty"])
            except (TypeError, ValueError):
                pass

        period = _join_distinct(r["period"] for r in group) or _parse_period(subject)

        merged_rows.append((
            ", ".join(days),
            period,
            dates[0] if dates else None,
            _join_distinct(r["labor_type"] for r in group),
            _join_distinct(r["time_type"] for r in group),
            total_qty,
            project_number,
            _join_distinct(r["Project Name"] for r in group),
            task_name,
            _join_distinct(r["name"] for r in group),
            _join_distinct(r["person_number"] for r in group),
            subject,
            sender,
            _join_distinct(r["received"] for r in group),
        ))

    conn.execute("DELETE FROM timecards_summary")
    conn.executemany("""
        INSERT INTO timecards_summary
        (day, "Period", "Date", labor_type, time_type, "Qty", "Project Number", "Project Name", "Task Name", "Name", "Person Number", subject, sender, received)
        VALUES
        (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, merged_rows)


def _sync_invoice_lines(conn):
    """
    Adds one invoice_lines row per raw timecards_approved entry (invoicing
    is approved-only), for any entry not already linked (via timecard_id).
    Only ever INSERTs new rows (OR IGNORE on the timecard_id UNIQUE key)
    -- an existing row, including any Invoice Number/Sales Price/etc
    you've filled in by hand, is never touched or overwritten by a later
    sync. "Date" is the day the email itself was received (not the
    timecard's work-day).
    """
    rows = conn.execute("""
        SELECT id, "Project Name", period, "Task Name", "Project Number", "Qty", received
        FROM timecards_approved
    """).fetchall()

    prepared = [
        (timecard_id, received[:10] if received else None, project_name, period, task_name, project_number, qty)
        for timecard_id, project_name, period, task_name, project_number, qty, received in rows
    ]

    conn.executemany("""
        INSERT OR IGNORE INTO invoice_lines
        (timecard_id, "Date", "Project Name", "Period", "Task Name", "Project Number", "Qty")
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, prepared)


def _prune_invoice_lines(conn):
    """
    Drops invoice_lines whose timecard is no longer in timecards_approved.
    An entry whose approval was revoked (it moved to Pending/Rejected -- see
    _save_row) isn't billable, and leaving its line behind would both invoice
    it anyway and double-bill it if it were approved again later: the
    re-approval inserts a fresh timecards_approved row, so _sync_invoice_lines
    would add a SECOND line beside the orphan.
    """
    conn.execute("""
        DELETE FROM invoice_lines
        WHERE timecard_id IS NOT NULL
          AND timecard_id NOT IN (SELECT id FROM timecards_approved)
    """)


def _dedupe_latest_across_statuses(grouped):
    """
    Collapses entries sharing the same _row_key down to the single
    latest-received one, ACROSS the status tables they'd otherwise be
    routed to. Takes and returns a {table_name: [row, ...]} mapping.

    Within one table, a batch can contain the same logical entry twice (a
    resend alongside the real one) and the last one written would win --
    but emails are scanned newest-received-first (see filter_service), so
    an OLDER duplicate lands later in the batch and would overwrite the
    newer one, exactly backwards. Across tables, one scan can pick up both
    a Pending email and the Approval that superseded it; only the later of
    the two is the entry's current state, and letting both through would
    leave a stale row in the other table.
    """
    best = {}
    for table_name, rows in grouped.items():
        for row in rows:
            key = _row_key(row)
            existing = best.get(key)
            if existing is None:
                best[key] = (table_name, row)
                continue
            new_parsed = _parse_received(row[13])
            existing_parsed = _parse_received(existing[1][13])
            if new_parsed and (existing_parsed is None or new_parsed > existing_parsed):
                best[key] = (table_name, row)

    resolved = {}
    for table_name, row in best.values():
        resolved.setdefault(table_name, []).append(row)
    return resolved


def _find_existing(conn, key):
    """
    Every stored row matching key, in whichever status table it sits.
    Returns (table_name, id, received, rate, is_exported, origin) tuples.
    """
    found = []
    for table_name in STATUS_TABLES.values():
        for row in conn.execute(f"""
            SELECT id, received, rate, is_exported, origin FROM "{table_name}"
            WHERE day = ? AND "Project Number" = ? AND "Task Name" = ?
              AND person_number = ? AND received_month = ?
        """, key):
            found.append((table_name, *row))
    return found


def _upsert_timecard(conn, table_name, row, origin=None):
    # The ON CONFLICT column list must name exactly the columns of the
    # table's UNIQUE constraint (see _create_timecards_table) -- SQLite
    # raises "ON CONFLICT clause does not match any PRIMARY KEY or UNIQUE
    # constraint" otherwise. The key columns are not in the DO UPDATE list
    # (they're what was matched on); sender is, since it's no longer part of
    # the key and a later forward of the same timecard should refresh it.
    # origin is also deliberately NOT in the DO UPDATE list, same reasoning
    # as rate/is_exported below -- who an entry originally came from isn't
    # a property of its latest received timestamp, so a later rescan or
    # resync must never flip it.
    conn.execute(f"""
        INSERT INTO "{table_name}"
        (day, "Date", labor_type, time_type, "Qty", "Project Number", "Project Name", "Task Name", name, period, person_number, subject, sender, received, received_month, origin)
        VALUES
        (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(day, "Project Number", "Task Name", person_number, received_month) DO UPDATE SET
            "Date" = excluded."Date",
            labor_type = excluded.labor_type,
            time_type = excluded.time_type,
            "Qty" = excluded."Qty",
            "Project Name" = excluded."Project Name",
            name = excluded.name,
            period = excluded.period,
            subject = excluded.subject,
            sender = excluded.sender,
            received = excluded.received
    """, (*row, origin))


def _save_row(conn, table_name, row, origin=None):
    """
    Writes one entry into its status table, holding the rule that an entry
    exists exactly once across the three of them: same day, project, task,
    person_number and received-MONTH is the same entry, and the latest
    "received" is its current state.

    So an entry that changes status within a month -- Pending or Rejected
    -> Approved, an approval later revoked, any direction -- is MOVED: the
    row is deleted from the table it used to sit in and written to the new
    one. An entry received in a DIFFERENT month is a different entry and
    simply coexists; the month is part of the key, so it never collides.

    A row already stored with a LATER "received" than the incoming one is
    left alone and the incoming row dropped -- re-reading an old Pending
    email must not undo the Approval that came after it.

    origin: which device this entry is attributed to. Only used to seed a
    brand-new row, or to backfill a legacy row that predates this column
    -- an entry that already has an origin keeps it regardless of who
    happens to be writing this particular update (see _upsert_timecard).
    """
    key = _row_key(row)
    incoming = _parse_received(row[13])
    existing = _find_existing(conn, key)

    for _table, _id, stored_received, _rate, _is_exported, _origin in existing:
        stored = _parse_received(stored_received)
        if stored is not None and (incoming is None or stored > incoming):
            return

    # A rate is typed in by hand, is_exported records that the entry went
    # out in a file, and origin records who it's attributed to; none of the
    # three is a property of the status it happened to be sitting under, so
    # they all ride along when the entry is moved.
    carried_rate = 0.0
    carried_is_exported = 0
    carried_origin = None
    for other_table, row_id, _received, rate, is_exported, existing_origin in existing:
        if other_table == table_name:
            continue  # same table: the upsert below updates it in place
        carried_rate = carried_rate or (rate or 0)
        carried_is_exported = max(carried_is_exported, is_exported or 0)
        carried_origin = carried_origin or existing_origin
        conn.execute(f'DELETE FROM "{other_table}" WHERE id = ?', (row_id,))

    _upsert_timecard(conn, table_name, row, origin=carried_origin or origin)

    if carried_rate or carried_is_exported or origin:
        conn.execute(f"""
            UPDATE "{table_name}" SET
                rate = CASE WHEN COALESCE(rate, 0) = 0 THEN ? ELSE rate END,
                is_exported = CASE WHEN COALESCE(is_exported, 0) = 0 THEN ? ELSE is_exported END,
                origin = COALESCE(origin, ?)
            WHERE day = ? AND "Project Number" = ? AND "Task Name" = ?
              AND person_number = ? AND received_month = ?
        """, (carried_rate, carried_is_exported, origin, *key))


def _group_by_status_table(entries):
    """
    Buckets entries by their target table (timecards_approved/_pending/
    _rejected) based on entry["status"]. Entries with a missing or
    unrecognized status are dropped -- there's no table to route them to.
    """
    grouped = {}
    skipped = 0
    for entry in entries:
        table_name = STATUS_TABLES.get(entry.get("status"))
        if table_name is None:
            skipped += 1
            continue
        grouped.setdefault(table_name, []).append(_to_row(entry))
    if skipped:
        print(f"Skipped {skipped} entr{'y' if skipped == 1 else 'ies'} with unrecognized/missing status.")
    return grouped


def save_card(entry: dict, origin=None):
    save_cards([entry], origin=origin)


def save_cards(entries: list, origin=None):
    """
    origin: which device these entries are attributed to. Defaults to
    THIS device's own id -- a real Outlook scan is always "mine", which
    is what lets build_outgoing_snapshot (see services/sharepoint_service.py)
    only ever publish rows this device actually scanned itself.
    """
    if origin is None:
        origin = get_device_id()

    grouped = _dedupe_latest_across_statuses(_group_by_status_table(entries))
    if not grouped:
        return

    conn = sqlite3.connect(DB_PATH)
    try:
        # Row by row rather than executemany: each one has to be weighed
        # against what's already stored under its key -- possibly in a
        # different status table -- before it can be written. See _save_row.
        for table_name, rows in grouped.items():
            for row in rows:
                _save_row(conn, table_name, row, origin=origin)
        _rebuild_summary(conn)
        _rebuild_project_type_tables(conn)
        _sync_invoice_lines(conn)
        # Comment out the next line to KEEP the invoice line of an entry whose
        # approval was later revoked (it moved to Pending/Rejected), instead of
        # deleting it. The trade either way:
        #   pruning (current)       -- the line goes, and any Invoice Number /
        #                              PO / Sales Price typed in by hand goes too.
        #   keeping (commented out) -- that manual data survives, but the entry
        #                              is invoiced despite no longer being
        #                              approved, and if it's ever approved again
        #                              it is billed TWICE: the re-approval is a
        #                              new timecards_approved row, so
        #                              _sync_invoice_lines adds a second line
        #                              next to the kept one.
        _prune_invoice_lines(conn)
        conn.commit()
    finally:
        # A raise anywhere above (a bad row, a schema mismatch) would otherwise
        # leak the connection and leave the database file locked for the rest
        # of the process -- the sync runs on a QThread, so the UI would keep
        # hitting a locked db long after the scan died.
        conn.close()


def _expense_amount(value):
    """Best-effort float for an amount string like "229.98" or "37,909".
    Returns None when it can't be parsed, so a bad cell doesn't sink the row."""
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


# A timecard record's "period" ("01/01/2026 - 01/07/2026") is what a
# report's expense_date is now checked against -- not a specific day
# anymore (see _fill_sender_timecard_match).
_period_range_pattern = re.compile(
    r"(?P<sm>\d{1,2})/(?P<sd>\d{1,2})/(?P<sy>\d{2,4})\s*-\s*"
    r"(?P<em>\d{1,2})/(?P<ed>\d{1,2})/(?P<ey>\d{2,4})"
)


def _full_year(year_str):
    year = int(year_str)
    return year if year > 99 else 2000 + year


def _period_bounds(period_text):
    """('01/01/2026 - 01/07/2026') -> (date(2026,1,1), date(2026,1,7)), or
    (None, None) if period_text doesn't parse."""
    match = _period_range_pattern.search(period_text or "")
    if not match:
        return None, None
    try:
        start = date(_full_year(match.group("sy")), int(match.group("sm")), int(match.group("sd")))
        end = date(_full_year(match.group("ey")), int(match.group("em")), int(match.group("ed")))
        return start, end
    except ValueError:
        return None, None


def _fetch_live_usd_rates():
    """Today's {currency_code: units-per-USD} table from a free, keyless
    FX API. Returns None on any failure (offline, API down, bad response)
    -- callers fall back to whatever's cached rather than blowing up a
    scan/export over a network hiccup."""
    try:
        with urllib.request.urlopen(_FX_API_URL, timeout=_FX_REQUEST_TIMEOUT_SECONDS) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        if payload.get("result") != "success" or not payload.get("rates"):
            return None
        return payload["rates"]
    except (urllib.error.URLError, TimeoutError, ValueError, OSError) as exc:
        print(f"[FX] Could not fetch live exchange rates: {exc}")
        return None


def _cached_usd_rates(conn):
    """Live {currency: units-per-USD} table, cached in app_state for
    _FX_CACHE_TTL_SECONDS -- so a whole batch of expenses scanned together
    shares one network round trip instead of each triggering its own.
    Falls back to whatever was last cached (however stale) when a fresh
    fetch fails; returns None only when there has never been a successful
    fetch at all.
    """
    rates_row = conn.execute(
        "SELECT value FROM app_state WHERE key = 'fx_rates_usd'"
    ).fetchone()
    cached = json.loads(rates_row[0]) if rates_row and rates_row[0] else None

    fetched_row = conn.execute(
        "SELECT value FROM app_state WHERE key = 'fx_rates_fetched_at'"
    ).fetchone()
    fetched_at = _parse_received(fetched_row[0]) if fetched_row and fetched_row[0] else None
    is_fresh = (
        fetched_at is not None
        and (datetime.now() - fetched_at).total_seconds() < _FX_CACHE_TTL_SECONDS
    )
    if cached is not None and is_fresh:
        return cached

    fresh = _fetch_live_usd_rates()
    if fresh is None:
        return cached  # stale-but-better-than-nothing, or None if never fetched

    conn.execute(
        "INSERT INTO app_state (key, value) VALUES ('fx_rates_usd', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (json.dumps(fresh),),
    )
    conn.execute(
        "INSERT INTO app_state (key, value) VALUES ('fx_rates_fetched_at', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),),
    )
    return fresh


def _to_usd(conn, amount, currency):
    """`amount` (in `currency`) converted to USD using the live/cached rate
    table, or None when it can't be converted (unrecognized/missing
    currency code, or no rate table has ever been fetched successfully)."""
    if amount is None:
        return None
    code = (currency or "").strip().upper()
    if code == "USD":
        return amount
    rates = _cached_usd_rates(conn)
    if not rates or not rates.get(code):
        return None
    return amount / rates[code]


def _expense_to_row(entry: dict) -> tuple:
    return (
        entry.get("expense_report"),
        entry.get("project_code"),
        entry.get("project_name"),
        _expense_amount(entry.get("report_total")),
        entry.get("report_currency"),
        entry.get("status"),
        entry.get("submitted_by"),
        entry.get("subject"),
        entry.get("sender"),
        entry.get("received"),
        _received_month(entry.get("received")),
        entry.get("link_method"),
        entry.get("matched_timecard_id"),
        entry.get("matched_sender"),
        entry.get("matched_person_number"),
        entry.get("matched_name"),
        entry.get("matched_period"),
        entry.get("matched_subject"),
        entry.get("matched_day"),
        entry.get("matched_project_number"),
        entry.get("matched_project_name"),
        entry.get("matched_task_name"),
        entry.get("amount_usd"),
        entry.get("expense_date"),
        entry.get("expense_date_start"),
        entry.get("matched_date"),
        entry.get("matched_received"),
    )


def _fill_sender_timecard_match(conn, entry):
    """
    Fills in an expense entry's matched_* fields by linking it to a whole
    timecard RECORD submitted by the same sender -- not to a specific day
    within one, which is how this used to work (see git history for the
    old day/date-based version).

    A "record" here is one submitted timecard: the group of per-day rows in
    timecards_approved that share the same (sender, period, subject,
    received) -- those four fields are identical across every day-row
    extracted out of a single approved-timecard email, so grouping on them
    recovers "one submission" out of the day-level table without needing a
    separate table for it.

    Preference order, among the sender's own records:
      1. A record whose period ("MM/DD/YYYY - MM/DD/YYYY") fully contains
         this report's own span, [expense_date_start, expense_date] (a
         report listing several Expense Item dates, e.g. a multi-day trip,
         must have ALL of those dates -- not just the latest one -- fall
         inside a single timecard period to count; a report with only one
         date collapses to a single-day span). If more than one record's
         period contains that span (overlapping periods), the
         most-recently-received one wins.
      2. If none of the sender's records has a period containing that full
         span (including when the span itself is unknown), fall back to
         that sender's single most-recently-received record.
      3. If the sender has no timecard record at all, link_method is
         "none" and nothing else is filled.

    Mutates entry in place. Sets link_method to "period_match" or
    "latest_fallback" accordingly. Day-specific fields that made sense for
    the old per-day match -- matched_timecard_id, matched_day, matched_date,
    matched_project_number, matched_project_name, matched_task_name -- have
    no single value at record level (a record can span several days,
    projects and tasks) and are simply left unset (NULL) here.
    """
    sender = entry.get("sender")
    if not sender:
        entry["link_method"] = "none"
        return

    records = conn.execute("""
        SELECT sender, period, subject, received, name, person_number
        FROM timecards_approved
        WHERE sender = ?
        GROUP BY sender, period, subject, received
    """, (sender,)).fetchall()

    if not records:
        entry["link_method"] = "none"
        return

    def _parse_iso_date(value):
        if not value:
            return None
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None

    expense_end = _parse_iso_date(entry.get("expense_date"))
    expense_start = _parse_iso_date(entry.get("expense_date_start"))
    if expense_start is None:
        expense_start = expense_end  # single-date report (or span unknown, if expense_end is also None)

    containing = []
    if expense_start is not None and expense_end is not None:
        for row in records:
            start, end = _period_bounds(row[1])
            if start and end and start <= expense_start and expense_end <= end:
                containing.append(row)

    pool = containing if containing else records
    link_method = "period_match" if containing else "latest_fallback"

    best = None
    best_received = None
    for row in pool:
        parsed_received = _parse_received(row[3])
        if best is None or (parsed_received is not None and (best_received is None or parsed_received > best_received)):
            best = row
            best_received = parsed_received

    tc_sender, period, subject, received, name, person_number = best
    entry["link_method"] = link_method
    entry["matched_sender"] = tc_sender
    entry["matched_person_number"] = person_number
    entry["matched_name"] = name
    entry["matched_period"] = period
    entry["matched_subject"] = subject
    entry["matched_received"] = received


def save_expenses(entries: list):
    """
    Persist expense reports (from extractor_service.extract_expense) -- one
    row per report. Keyed on expense_report: re-reading the same report
    refreshes its existing row in place rather than inserting a duplicate.

    Every entry is linked here to the timecard record submitted by the
    same sender (see _fill_sender_timecard_match) -- extract_expense no
    longer does any matching itself.

    Each entry also gets amount_usd filled in here: its own reported
    amount/currency converted with the live rate in effect right now (see
    _to_usd). Doing the conversion once, at save time, means it's frozen to
    the rate on the day the report was actually scanned rather than
    silently drifting every time an export re-reads the row later.
    """
    if not entries:
        return

    conn = sqlite3.connect(DB_PATH)
    try:
        _create_expenses_table(conn)
        for entry in entries:
            _fill_sender_timecard_match(conn, entry)
            entry["amount_usd"] = _to_usd(
                conn, _expense_amount(entry.get("report_total")), entry.get("report_currency")
            )
        conn.executemany("""
            INSERT INTO expenses
            (expense_report, "Project Number", "Project Name", "Amount", currency,
             status, submitted_by, subject, sender, received, received_month,
             link_method, matched_timecard_id, matched_sender, matched_person_number,
             matched_name, matched_period, matched_subject, matched_day,
             matched_project_number, matched_project_name, matched_task_name, amount_usd,
             expense_date, expense_date_start, matched_date, matched_received)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(expense_report) DO UPDATE SET
                "Project Number" = excluded."Project Number",
                "Project Name" = excluded."Project Name",
                "Amount" = excluded."Amount",
                currency = excluded.currency,
                status = excluded.status,
                submitted_by = excluded.submitted_by,
                subject = excluded.subject,
                sender = excluded.sender,
                received = excluded.received,
                received_month = excluded.received_month,
                link_method = excluded.link_method,
                matched_timecard_id = excluded.matched_timecard_id,
                matched_sender = excluded.matched_sender,
                matched_person_number = excluded.matched_person_number,
                matched_name = excluded.matched_name,
                matched_period = excluded.matched_period,
                matched_subject = excluded.matched_subject,
                matched_day = excluded.matched_day,
                matched_project_number = excluded.matched_project_number,
                matched_project_name = excluded.matched_project_name,
                matched_task_name = excluded.matched_task_name,
                amount_usd = excluded.amount_usd,
                matched_received = excluded.matched_received,
                expense_date = excluded.expense_date,
                expense_date_start = excluded.expense_date_start,
                matched_date = excluded.matched_date
        """, [_expense_to_row(e) for e in entries])
        conn.commit()
    finally:
        conn.close()


def get_expense_rows(start_date=None, end_date=None, project_type=None):
    """All expense line-item rows as dicts, newest report first. When a date
    range is given, only rows whose "received" falls within it are returned
    (same rule the timecard views use).

    project_type ("beverage"/"hospitality", or None for all) narrows the rows
    to one division by the same "Project Name" prefix rule the timecards use --
    an expense on project "HLGIU-..." lands under Hospitality, so the Dashboard
    can filter timecards and expenses with one shared control."""
    conn = sqlite3.connect(DB_PATH)
    try:
        if conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='expenses'"
        ).fetchone() is None:
            return []
        where, params = _project_type_clause(project_type)
        where_sql = f" WHERE {where}" if where else ""
        cursor = conn.execute(
            f'SELECT * FROM expenses{where_sql} ORDER BY received DESC, expense_report ASC',
            params,
        )
        columns = [desc[0] for desc in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        if start_date is not None or end_date is not None:
            rows = [r for r in rows if _received_in_range(r.get("received"), start_date, end_date)]
        return rows
    finally:
        conn.close()


def get_expense_count(start_date=None, end_date=None, project_type=None):
    """Number of expense line-item rows in the given window/division -- backs
    the Dashboard's Expenses stat card."""
    return len(get_expense_rows(start_date, end_date, project_type))


def get_expense_columns():
    """Column names of the expenses table, in schema order -- lets the grid
    show the right headings on an empty table (before any expenses exist)."""
    conn = sqlite3.connect(DB_PATH)
    try:
        if conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='expenses'"
        ).fetchone() is None:
            return []
        return [row[1] for row in conn.execute('PRAGMA table_info("expenses")')]
    finally:
        conn.close()


def export_expenses_to_csv(output_path="expenses.csv"):
    """Dump the expenses table to a CSV, one row per report."""
    conn = sqlite3.connect(DB_PATH)
    try:
        cursor = conn.execute('SELECT * FROM expenses ORDER BY received ASC, expense_report ASC')
        columns = [desc[0] for desc in cursor.description]
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            writer.writerows(cursor.fetchall())
        _record_export(conn, os.path.basename(output_path))
        conn.commit()
    finally:
        conn.close()
    print(f"Exported expenses to {output_path}")


def _row_dict_to_entry(row_dict, status_label):
    """
    Converts one stored row (a dict straight off SELECT *) back into the
    same "entry" shape extract() produces -- the shape _to_row()/
    save_cards() already know how to consume. Used by
    build_outgoing_snapshot to hand the SharePoint file channel (see
    services/sharepoint_service.py) rows in the same shape a real Outlook
    scan produces, instead of a second parallel shape that could drift
    out of sync with the real one over time.

    rate/rate_updated_at/rate_updated_by ride along as extra keys that
    _to_row() itself ignores -- they're just carried through to the
    rendered current-sheet xlsx for reference.
    """
    return {
        "day": row_dict.get("day"),
        "received": row_dict.get("received"),
        "labor_type": row_dict.get("labor_type"),
        "time_type": row_dict.get("time_type"),
        "hours": row_dict.get("Qty"),
        "project_code": row_dict.get("Project Number"),
        "project_name": row_dict.get("Project Name"),
        "task": row_dict.get("Task Name"),
        "name": row_dict.get("name"),
        "period": row_dict.get("period"),
        "person_number": row_dict.get("person_number"),
        "subject": row_dict.get("subject"),
        "sender": row_dict.get("sender"),
        "status": status_label,
        "rate": row_dict.get("rate"),
        "rate_updated_at": row_dict.get("rate_updated_at"),
        "rate_updated_by": row_dict.get("rate_updated_by"),
    }


_UNSET = object()  # distinguishes "since_date not passed" from "since_date=None" below


def build_outgoing_snapshot(project_type=None, since_date=_UNSET):
    """
    Packages everything THIS device has scanned itself (origin = this
    device's id -- see save_cards) for the current open period into a
    dict ready to be mailed to the other user and handed to their
    apply_incoming_snapshot() -- or, via the SharePoint file channel (see
    services/sharepoint_service.py), rendered to this device's own
    current_<device_id>.xlsx.

    since_date controls where the open period starts:
      - omitted entirely -> defaults to get_last_export_date() (the email
        sync channel's boundary), so that flow is unchanged.
      - explicitly passed (including None, meaning "no lower bound -- the
        beginning of time") -> used as-is. This is how the SharePoint
        channel passes its own, independent boundary.json boundary_date
        instead of the email channel's last_export_date -- the two
        boundaries must stay separate.

    Deliberately excludes any row this device only knows about because it
    was itself synced in from the OTHER device -- each device only ever
    publishes what it originally scanned. The other side already has
    anything it sent; re-sending it back would just make every payload
    grow forever and teach the other side nothing new.

    "seq" in the returned dict is a monotonically increasing per-device
    outgoing counter (see _next_outgoing_seq) -- only meaningful to the
    email channel's apply_incoming_snapshot supersede check; the
    SharePoint channel ignores it entirely.
    """
    device_id = get_device_id()
    period_start = get_last_export_date() if since_date is _UNSET else since_date
    type_where, type_params = _project_type_clause(project_type)

    conn = sqlite3.connect(DB_PATH)
    try:
        rows = []
        for status_label, table_name in STATUS_TABLES.items():
            clauses = ["origin = ?"]
            params = [device_id]
            if period_start:
                clauses.append("received >= ?")
                params.append(period_start)
            if type_where:
                clauses.append(type_where)
                params.extend(type_params)

            cursor = conn.execute(
                f'SELECT * FROM "{table_name}" WHERE {" AND ".join(clauses)}', params
            )
            columns = [desc[0] for desc in cursor.description]
            for raw in cursor.fetchall():
                rows.append(_row_dict_to_entry(dict(zip(columns, raw)), status_label))

        seq = _next_outgoing_seq(conn)
        conn.commit()
    finally:
        conn.close()

    return {
        "device_id": device_id,
        "seq": seq,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "period_start": period_start,
        "rows": rows,
    }


def apply_incoming_snapshot(payload):
    """
    Merges another device's build_outgoing_snapshot() payload into this
    database. Safe to call more than once with the same payload, and safe
    if payloads from the same sender arrive out of order -- it's only
    actually applied when its seq is strictly newer than the last one
    already applied FROM THAT SENDER (see get_last_applied_seq); an
    older or repeated one is silently skipped rather than reprocessed.

    That's a deliberate design choice, not just deduplication: because
    each snapshot is the sender's FULL current-period picture rather than
    a delta, the newest one queued from a given sender is always a
    superset of any older one still sitting unprocessed -- so if several
    sync mails from the same device pile up (e.g. this laptop was closed
    for a few days), applying only the newest and discarding the rest is
    correct, not lossy.

    Returns a dict describing what happened, for the caller (the email
    import step) to log/surface to the user.
    """
    if not payload:
        return {"applied": False, "reason": "empty payload"}

    device_id = payload.get("device_id")
    incoming_seq = payload.get("seq") or 0
    if not device_id:
        return {"applied": False, "reason": "missing device_id"}

    last_applied = get_last_applied_seq(device_id)
    if incoming_seq <= last_applied:
        return {
            "applied": False, "reason": "superseded",
            "seq": incoming_seq, "last_applied": last_applied,
        }

    rows = payload.get("rows") or []

    # The timecard data itself (day/project/task/status/etc) goes through
    # the exact same path a real scan uses -- see save_cards -- just
    # tagged with the SENDER's device id as origin instead of ours, which
    # is what keeps this device from ever re-broadcasting it back out.
    save_cards(rows, origin=device_id)

    conn = sqlite3.connect(DB_PATH)
    try:
        rates_applied = 0
        for entry in rows:
            if not entry.get("rate_updated_at"):
                continue
            if _apply_rate_if_newer(
                conn,
                status_label=entry.get("status"),
                natural_key=_row_key(_to_row(entry)),
                rate=entry.get("rate"),
                updated_at=entry.get("rate_updated_at"),
                updated_by=entry.get("rate_updated_by"),
            ):
                rates_applied += 1

        _set_last_applied_seq(conn, device_id, incoming_seq)
        conn.commit()
    finally:
        conn.close()

    return {
        "applied": True, "seq": incoming_seq,
        "rows_merged": len(rows), "rates_applied": rates_applied,
    }


def _apply_rate_if_newer(conn, status_label, natural_key, rate, updated_at, updated_by):
    """
    The shared last-write-wins merge behind BOTH the per-row rate carried
    inside a snapshot and a standalone rate-update message (see
    build_rate_update_payload/apply_rate_update below). Only overwrites
    the locally stored rate if the incoming edit is strictly newer than
    whatever timestamp is already stored -- an incoming edit that's the
    same age or older than the local one is left alone, so replaying an
    old message (or two devices' clocks disagreeing by a few seconds)
    can't undo a genuinely newer local edit.
    """
    table_name = STATUS_TABLES.get(status_label)
    if table_name is None or rate is None or not updated_at:
        return False

    row = conn.execute(
        f'SELECT id, rate_updated_at FROM "{table_name}" '
        f'WHERE day = ? AND "Project Number" = ? AND "Task Name" = ? '
        f'AND person_number = ? AND received_month = ?',
        natural_key,
    ).fetchone()
    if row is None:
        return False  # the timecard itself hasn't landed here -- nothing to attach a rate to

    record_id, stored_updated_at = row
    incoming_parsed = _parse_received(updated_at)
    stored_parsed = _parse_received(stored_updated_at) if stored_updated_at else None
    if incoming_parsed is not None and stored_parsed is not None and stored_parsed >= incoming_parsed:
        return False

    conn.execute(
        f'UPDATE "{table_name}" SET rate = ?, rate_updated_at = ?, rate_updated_by = ? WHERE id = ?',
        (rate, updated_at, updated_by, record_id),
    )
    return True


def build_rate_update_payload(status_key, record_id):
    """
    A small standalone message for ONE rate edit, meant to be sent right
    after the edit happens rather than waiting for the next full scan/
    snapshot -- the point is the other user seeing a rate you just typed
    in without needing to Scan Inbox first. Returns None if record_id
    doesn't exist.
    """
    table_name = STATUS_TABLES.get(_STATUS_LABELS.get(status_key, "Approved"))
    if table_name is None:
        return None

    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            f'SELECT day, "Project Number", "Task Name", person_number, received_month, '
            f'rate, rate_updated_at, rate_updated_by FROM "{table_name}" WHERE id = ?',
            (record_id,),
        ).fetchone()
    finally:
        conn.close()
    if row is None:
        return None

    day, project_number, task_name, person_number, received_month, rate, rate_updated_at, rate_updated_by = row
    return {
        "device_id": get_device_id(),
        "status": _STATUS_LABELS.get(status_key, "Approved"),
        "natural_key": [day, project_number, task_name, person_number, received_month],
        "rate": rate,
        "rate_updated_at": rate_updated_at,
        "rate_updated_by": rate_updated_by,
    }


def apply_rate_update(payload):
    """Applies one standalone rate-update message (see
    build_rate_update_payload). Returns True if it actually changed
    anything (i.e. it was newer than what's already stored)."""
    if not payload:
        return False
    natural_key = tuple(payload.get("natural_key") or ())
    if len(natural_key) != 5:
        return False

    conn = sqlite3.connect(DB_PATH)
    try:
        applied = _apply_rate_if_newer(
            conn,
            status_label=payload.get("status"),
            natural_key=natural_key,
            rate=payload.get("rate"),
            updated_at=payload.get("rate_updated_at"),
            updated_by=payload.get("rate_updated_by"),
        )
        conn.commit()
    finally:
        conn.close()
    return applied


def record_finalize_from_other_device(name, end_date):
    """
    Applied when THIS device receives a "finalize" sync message from the
    other user (see sync_service.pull_updates) -- logs their export into
    THIS device's own export_history and advances THIS device's
    last_export_date to match theirs.

    This is what makes the finalize boundary SHARED state instead of
    something each machine tracks independently: without it, the device
    that didn't click Finalize would keep treating the just-closed period
    as still open, and any timecard received in the last few days of it
    would get scanned into what it thinks is next month's data instead.
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        _record_export(conn, name)
        _set_last_export_date(conn, end_date)
        conn.commit()
    finally:
        conn.close()


def export_to_csv(output_path="output.csv"):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute('SELECT * FROM timecards_summary ORDER BY "Date" ASC')
    columns = [desc[0] for desc in cursor.description]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(cursor.fetchall())

    conn.close()
    print(f"Exported to {output_path}")


def export_summary_csv_range(start_date: str, end_date: str, output_path: str, project_type=None) -> int:
    """
    Exports approved timecard rows whose "Date" falls within
    [start_date, end_date] (inclusive, both 'YYYY-MM-DD' strings) to a
    CSV at output_path. Backs the History page's "Export last month" /
    "Export date range" buttons. Records the export in export_history,
    same as the other export functions, so it shows up in that list too.
    Returns the number of rows written.

    project_type ("beverage"/"hospitality", or None for every project)
    narrows the export to one division, by the same project-name rule the
    project-type tables are built on -- so an export can cover just Food &
    Beverage or just Hospitality. Only the rows actually written are
    flagged is_exported, so exporting one division leaves the other's
    entries still marked un-exported.

    Reads timecards_approved rather than timecards_summary so the export
    keeps one row per day, exactly like the Dashboard's records table
    (see get_status_rows) -- the summary table merges a whole timecard
    email into a single row with the days joined and Qty summed. The
    leading columns are the Dashboard's, in its order; the rest follow so
    no field is dropped.

    Every exported record gets its is_exported flag set to 1 (a no-op for
    rows already flagged from an earlier export). The id is fetched only
    for that flagging and is not written to the CSV; is_exported itself is
    left out of the file too -- it's bookkeeping, not timecard data.
    """
    type_where, type_params = _project_type_clause(project_type)
    type_sql = f" AND {type_where}" if type_where else ""

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute(
        'SELECT id, subject, "Project Number", "Project Name", "Task Name", "Date", "Qty", rate, '
        'day, period, labor_type, time_type, name, person_number, sender, received '
        f'FROM timecards_approved WHERE "Date" >= ? AND "Date" <= ?{type_sql} '
        'ORDER BY "Date" ASC, subject ASC',
        [start_date, end_date, *type_params],
    )
    columns = [desc[0] for desc in cursor.description][1:]  # drop id
    rows = cursor.fetchall()

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(row[1:] for row in rows)

    conn.executemany(
        "UPDATE timecards_approved SET is_exported = 1 WHERE id = ? AND is_exported != 1",
        [(row[0],) for row in rows],
    )
    _record_export(conn, os.path.basename(output_path))
    # Remember how far this export reached, so the next one can start where
    # this one left off. Overwritten every export (see _set_last_export_date).
    _set_last_export_date(conn, end_date)
    conn.commit()
    conn.close()
    print(f"Exported {len(rows)} row(s) ({start_date} to {end_date}) to {output_path}")
    return len(rows)


def _linked_expenses_by_timecard(conn, timecard_ids):
    """
    {timecard_row_id: total_linked_expense_amount_usd} for the given
    timecards_approved ids, out of every expense linked to one of them (see
    storage_service._fill_sender_timecard_match for how that link is made).

    Linking now happens at the RECORD level (sender+period+subject+received
    -- one submitted timecard, which is usually several day-rows), not to
    one specific day-row. To resolve a linked expense back to real ids here:
    every row in timecard_ids is grouped by that same (sender, period,
    subject, received) key, each expense (link_method != 'none') is matched
    to the group sharing its matched_sender/matched_period/matched_subject/
    matched_received, and that expense's amount_usd is added to the group's
    running total.

    Each record's total is then attributed to a single representative
    day-row -- the lowest id in the group -- with every other day-row in
    that same record left out of the returned dict entirely. This is
    deliberate: export_act_invoice_overview_range writes one Expense row
    per day-row and SUMs them for the invoice total, so putting the same
    record-level total on every one of its days would multiply it by
    however many days that timecard covered.

    Sums amount_usd, not the report's own raw "Amount" -- two reports on
    the same record aren't necessarily in the same currency, so adding
    their face-value amounts together would be meaningless. amount_usd is
    already each report's own amount converted at the rate in effect when
    it was scanned (see _to_usd), so summing it is a real total. A record
    with more than one linked expense therefore gets both added together
    here rather than one silently overwriting the other -- and a report
    that couldn't be converted (amount_usd is NULL) contributes nothing
    rather than sinking the whole sum.
    """
    if not timecard_ids:
        return {}

    placeholders = ",".join("?" * len(timecard_ids))
    rows = conn.execute(f"""
        SELECT id, sender, period, subject, received FROM timecards_approved
        WHERE id IN ({placeholders})
    """, timecard_ids).fetchall()

    # One representative id per (sender, period, subject, received) group --
    # the lowest id in that group -- to attribute the whole record's linked
    # total to a single day-row rather than every day-row in it.
    representative_id = {}
    for tc_id, sender, period, subject, received in rows:
        key = (sender, period, subject, received)
        if key not in representative_id or tc_id < representative_id[key]:
            representative_id[key] = tc_id

    linked = conn.execute("""
        SELECT matched_sender, matched_period, matched_subject, matched_received, amount_usd
        FROM expenses WHERE link_method != 'none'
    """).fetchall()

    totals = {}
    for m_sender, m_period, m_subject, m_received, amount_usd in linked:
        key = (m_sender, m_period, m_subject, m_received)
        tc_id = representative_id.get(key)
        if tc_id is not None:
            totals[tc_id] = (totals.get(tc_id) or 0) + (amount_usd or 0)

    return totals


def export_act_invoice_overview_range(start_date: str, end_date: str, output_path: str, project_type=None) -> int:
    """
    Exports approved timecard rows whose "Date" falls within
    [start_date, end_date] straight into the ACT "Invoice Overview per
    Period" template layout (.xlsx) -- same query/filters/bookkeeping as
    export_summary_csv_range, but the file this produces opens as the
    proper template instead of a flat CSV.

    Each timecard entry becomes a LABOR row immediately followed by an
    Expense row, mirroring the template's existing pattern:
      LABOR row   -> Date, PO (fixed), Line=1, Type=LABOR, Project Number,
                     Project Name, Period, Task Name, Qty, Sales Price
                     (=rate, converted from AED to USD at the live rate --
                     see below), Total Amount (=Qty*Sales Price formula).
                     Invoice No / Project Mgr / SOW / Consultant are left
                     blank -- not tracked by this system.
      Expense row -> Line=2, Type=Expense, and the Total Amount formula.
                     Sales Price is that record's own linked expense
                     amount (already in USD -- see
                     _linked_expenses_by_timecard) when one exists, or 0
                     -- an unchanged blank row -- otherwise.

    Every monetary figure on the sheet -- LABOR, Expense, and the
    LABOR/EXPENSE/Invoice Total/VAT totals block -- ends up in USD, so
    nothing on it silently mixes currencies. rate is entered by hand
    elsewhere in AED (that's the one implicit currency this whole system
    otherwise assumes), so it's converted here with the same live/cached
    rate table _to_usd uses for expenses. If no AED rate can be resolved
    at all (never fetched successfully, and offline right now) this
    raises rather than exporting an invoice with silently wrong LABOR
    totals -- that's worse than an export simply failing.

    Returns the number of source timecard rows written (i.e. half the
    number of spreadsheet rows, since each becomes a LABOR+Expense pair).
    """
    type_where, type_params = _project_type_clause(project_type)
    type_sql = f" AND {type_where}" if type_where else ""

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute(
        f'SELECT {_ACT_ROW_COLUMNS} '
        f'FROM timecards_approved WHERE "Date" >= ? AND "Date" <= ?{type_sql} '
        'ORDER BY "Date" ASC, subject ASC',
        [start_date, end_date, *type_params],
    )
    rows = cursor.fetchall()

    try:
        _write_act_invoice_workbook(conn, rows, output_path)
    except Exception:
        conn.close()
        raise

    conn.executemany(
        "UPDATE timecards_approved SET is_exported = 1 WHERE id = ? AND is_exported != 1",
        [(row[0],) for row in rows],
    )
    _record_export(conn, os.path.basename(output_path))
    _set_last_export_date(conn, end_date)
    conn.commit()
    conn.close()
    print(f"Exported {len(rows)} row(s) ({start_date} to {end_date}) to {output_path}")
    return len(rows)


# The exact column list _write_act_invoice_workbook expects a row tuple to
# be, in order -- kept in one place so the range export and the rolling
# active export below can't drift apart on it.
_ACT_ROW_COLUMNS = 'id, "Project Number", "Project Name", "Task Name", "Qty", rate, day, period'


def _write_act_invoice_workbook(conn, rows, output_path):
    """
    Writes `rows` (tuples in _ACT_ROW_COLUMNS order) out as the ACT
    "Invoice Overview per Period" sheet at output_path, overwriting
    whatever is there. Pure file-writing: no is_exported flagging, no
    export_history entry, no commit -- the caller owns all of that, which
    is what lets the rolling active export (see rebuild_active_export)
    rewrite its file over and over without each rewrite counting as a
    separate export in the log.

    Takes an open conn only to read from it (linked expenses, cached FX).
    """
    linked_expenses = _linked_expenses_by_timecard(conn, [row[0] for row in rows])

    aed_per_usd = None
    if rows:
        fx_rates = _cached_usd_rates(conn)
        aed_per_usd = fx_rates.get("AED") if fx_rates else None
        if not aed_per_usd:
            raise RuntimeError(
                "Could not get a live AED -> USD exchange rate (no internet right "
                "now, and none cached from an earlier run) -- refusing to export an "
                "invoice with unconverted LABOR amounts. Try again once you have a "
                "connection."
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "Billed Invoice Details"

    ws.merge_cells("B2:D2")
    ws["B2"] = _ACT_TITLE

    header_row = 4
    for col_offset, header in enumerate(_ACT_HEADERS):
        cell = ws.cell(row=header_row, column=2 + col_offset, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    r = header_row + 1
    for _id, project_number, project_name, task_name, qty, rate, day, period in rows:
        qty_val = float(qty) if qty not in (None, "") else None
        rate_val = float(rate) if rate not in (None, "") else 0.0
        rate_usd = rate_val / aed_per_usd

        # LABOR row
        ws.cell(row=r, column=2, value=day)                 # Date (no year, as extracted)
        ws.cell(row=r, column=4, value=project_name)         # Project Name
        ws.cell(row=r, column=5, value=period)               # Period
        ws.cell(row=r, column=6, value=task_name)            # Task Name
        ws.cell(row=r, column=8, value=_ACT_PO)              # PO
        ws.cell(row=r, column=10, value=1)                   # Line
        ws.cell(row=r, column=11, value="LABOR")              # Type
        ws.cell(row=r, column=12, value=project_number)      # Project Number
        ws.cell(row=r, column=14, value=qty_val)              # Qty
        sp_cell = ws.cell(row=r, column=15, value=rate_usd)   # Sales Price (USD)
        sp_cell.number_format = _USD_FORMAT
        total_cell = ws.cell(row=r, column=16, value=f"=N{r}*O{r}")  # Total Amount
        total_cell.number_format = _USD_FORMAT
        r += 1

        # Expense row -- Line=2, Type, Total formula; Sales Price is this
        # record's own linked expense amount (already converted to USD --
        # see _linked_expenses_by_timecard) when it has one, 0 otherwise.
        ws.cell(row=r, column=10, value=2)
        ws.cell(row=r, column=11, value="Expense")
        sp_cell = ws.cell(row=r, column=15, value=linked_expenses.get(_id) or 0)
        sp_cell.number_format = _USD_FORMAT
        total_cell = ws.cell(row=r, column=16, value=f"=N{r}*O{r}")
        total_cell.number_format = _USD_FORMAT
        r += 1

    last_data_row = r - 1

    table = Table(displayName="Table4", ref=f"B{header_row}:P{last_data_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium23", showRowStripes=True,
    )
    ws.add_table(table)

    # Totals block, right after the data
    t = last_data_row + 1
    ws.cell(row=t, column=12, value="Total LABOR")
    ws.cell(row=t, column=14, value="=SUBTOTAL(109,Table4[Qty])")
    ws.cell(row=t, column=16, value='=SUMIF(Table4[Type],"LABOR",Table4[Total Amount])')

    ws.cell(row=t + 1, column=12, value="Total EXPENSE")
    ws.cell(row=t + 1, column=16, value='=SUMIF(Table4[Type],"Expense",Table4[Total Amount])')

    ws.cell(row=t + 2, column=12, value="Invoice Total ")
    ws.cell(row=t + 2, column=16, value=f"=P{t}+P{t + 1}")

    ws.cell(row=t + 3, column=13, value="VAT")
    ws.cell(row=t + 3, column=16, value=f"=P{t + 2}*0.05")

    ws.cell(row=t + 4, column=13, value="Total with VAT")
    ws.cell(row=t + 4, column=16, value=f"=P{t + 2}+P{t + 3}")

    for total_row in (t, t + 1, t + 2, t + 3, t + 4):
        cell = ws.cell(row=total_row, column=16)
        cell.number_format = _USD_FORMAT

    for col_letter, width in zip("BCDEFGHIJKLMNOP", [12, 11, 30, 16, 30, 13, 14, 8, 6, 8, 14, 14, 7, 13, 15]):
        ws.column_dimensions[col_letter].width = width

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)


# ----------------------------------------------------------------------
# The rolling "active" export -- the file the Update button keeps topping
# up, until Finalize closes it and points the app at a fresh one.
#
# Two pieces of state back this:
#   app_state['active_export_path']   -- the file currently being filled
#   active_export_rows                -- which timecards_approved ids are
#                                        already in that file
# Update adds only rows NOT in active_export_rows, then rewrites the whole
# file from the accumulated set (rewriting rather than appending in place:
# the sheet carries a Table ref and a totals block underneath the data, so
# every added row would have to shift and re-anchor those anyway -- a
# rewrite from the id set is the same result with nothing to keep in sync).
# Finalize leaves that file alone as the finished article, empties
# active_export_rows, and writes a new active_export_path -- so the next
# Update starts a fresh file instead of re-topping-up a closed one.
# ----------------------------------------------------------------------

# The one folder every .xlsx this app produces lands in, so the sheets
# aren't scattered next to output.csv and the source tree.
EXPORTS_DIR = os.path.join(BASE_DIR, "exports")


def _new_active_export_path():
    """
    Path for a newly opened active export: the current month, then a
    counter. 2026-07.xlsx is July's first sheet, 2026-07_2.xlsx the one
    opened after that was finalized, and so on -- so the name says which
    month's work it holds, and finalizing twice in a month can't land the
    new sheet on top of the one just closed.
    """
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    base = os.path.join(EXPORTS_DIR, date.today().strftime("%Y-%m"))
    path = f"{base}.xlsx"
    attempt = 2
    while os.path.exists(path):
        path = f"{base}_{attempt}.xlsx"
        attempt += 1
    return path


def get_active_export_path():
    """The file Update is currently filling, or None if no active export
    has been opened yet (fresh install, or straight after a finalize)."""
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            "SELECT value FROM app_state WHERE key = 'active_export_path'"
        ).fetchone()
        return row[0] if row and row[0] else None
    finally:
        conn.close()


def _set_active_export_path(conn, path):
    conn.execute(
        "INSERT INTO app_state (key, value) VALUES ('active_export_path', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (path,),
    )


def rebuild_active_export(project_type=None):
    """
    Behind the History page's Update button, and the export half of
    Finalize. Adds every approved timecard row that isn't in the active
    export yet, then rewrites the active file from the full accumulated
    set. Reads only the database -- scanning the inbox is Scan Inbox's
    job, and whatever it found is already stored by the time this runs.

    Opens a new active file if there isn't one (first ever Update, or the
    first Update after a Finalize) -- that's the "creates the excel the
    first time" case; every later call tops the same file up.

    Nothing here touches export_history or last_export_date: an in-progress
    file isn't an export yet. Finalize is what logs it (see
    finalize_active_export).

    Returns {"path", "new_rows", "total_rows", "created"}.
    """
    conn = sqlite3.connect(DB_PATH)
    try:
        row = conn.execute(
            "SELECT value FROM app_state WHERE key = 'active_export_path'"
        ).fetchone()
        path = row[0] if row and row[0] else None
        created = path is None
        if created:
            path = _new_active_export_path()
            _set_active_export_path(conn, path)

        type_where, type_params = _project_type_clause(project_type)
        type_sql = f" AND {type_where}" if type_where else ""

        # "New" means never written into ANY rolling sheet -- finalized
        # ones included, so a closed sheet's rows don't all pour back into
        # the fresh one right after a Finalize.
        new_ids = [
            r[0] for r in conn.execute(
                "SELECT id FROM timecards_approved WHERE id NOT IN "
                f"(SELECT timecard_id FROM active_export_rows){type_sql}",
                type_params,
            ).fetchall()
        ]
        conn.executemany(
            "INSERT OR IGNORE INTO active_export_rows "
            "(timecard_id, added_at, export_path, finalized) VALUES (?, ?, ?, 0)",
            [(new_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), path) for new_id in new_ids],
        )

        rows = conn.execute(
            f'SELECT {_ACT_ROW_COLUMNS} FROM timecards_approved '
            'WHERE id IN (SELECT timecard_id FROM active_export_rows WHERE finalized = 0) '
            'ORDER BY "Date" ASC, subject ASC'
        ).fetchall()

        # Written before the commit on purpose: if the sheet can't be
        # written (no FX rate, file open in Excel), the rows must NOT stay
        # marked as already-in-the-file, or the next Update would skip them.
        _write_act_invoice_workbook(conn, rows, path)
        conn.commit()
    finally:
        conn.close()

    print(f"Active export {path}: +{len(new_ids)} new row(s), {len(rows)} total")
    return {"path": path, "new_rows": len(new_ids), "total_rows": len(rows), "created": created}


def finalize_active_export(end_date, project_type=None):
    """
    Closes out the active export: one last top-up (so anything that
    arrived since the previous Update is in the file), then the file is
    logged in export_history, its rows are flagged is_exported, and the
    pointer is cleared so the NEXT Update opens a brand new file and
    starts filling that one instead.

    Returns {"path", "row_count"} -- path being the file just closed,
    which is what gets mailed to the other user.
    """
    result = rebuild_active_export(project_type=project_type)
    path = result["path"]

    conn = sqlite3.connect(DB_PATH)
    try:
        ids = [
            r[0] for r in
            conn.execute("SELECT timecard_id FROM active_export_rows WHERE finalized = 0").fetchall()
        ]
        conn.executemany(
            "UPDATE timecards_approved SET is_exported = 1 WHERE id = ? AND is_exported != 1",
            [(row_id,) for row_id in ids],
        )
        _record_export(conn, os.path.basename(path))
        _set_last_export_date(conn, end_date)
        # Rows marked done (not deleted -- see init_db) + no pointer =
        # the next Update opens a new file and fills it with new rows only.
        conn.execute("UPDATE active_export_rows SET finalized = 1 WHERE finalized = 0")
        conn.execute("DELETE FROM app_state WHERE key = 'active_export_path'")
        conn.commit()
    finally:
        conn.close()

    print(f"Finalized {path} with {result['total_rows']} row(s)")
    return {"path": path, "row_count": result["total_rows"]}


def export_invoice_lines_to_excel(output_path=None):
    """
    Exports invoice_lines to a real formatted .xlsx (not .csv - a CSV is
    plain text and can't hold fonts/colors/borders at all). Header row is
    bold, larger, light-blue-filled; every cell in the table gets a thin
    border and black text.

    Defaults into EXPORTS_DIR rather than the working directory, so every
    .xlsx this app writes ends up in the same folder.
    """
    if output_path is None:
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        output_path = os.path.join(EXPORTS_DIR, "invoice_lines.xlsx")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute("""
        SELECT "Date", "Invoice Number", "Project Name", "Period", "Task Name", "Project Mgr",
               "PO", "SOW", "Line", "Type", "Project Number", "Consultant", "Qty", "Sales Price",
               "Total Amount"
        FROM invoice_lines
        ORDER BY "Date" ASC
    """)
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Lines"

    ws.append(columns)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER

    for i, row in enumerate(rows):
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
        if i < len(rows) - 1:
            ws.append([])  # blank separator row between records - no styling

    for i, column_name in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(len(column_name) + 2, 12)

    wb.save(output_path)
    _record_export(conn, os.path.basename(output_path))
    conn.commit()
    conn.close()
    print(f"Exported to {output_path}")


# entry key -> display header, in write order. Read back by
# services/sharepoint_service.py (_read_device_sheet) to merge multiple
# devices' current sheets -- keep the two in sync; this is the one place
# the column layout is defined. rate_updated_at/rate_updated_by are
# included for reference/audit even though nothing currently merges on
# them.
SNAPSHOT_COLUMNS = [
    ("day", "Day"),
    ("received", "Received"),
    ("status", "Status"),
    ("project_code", "Project Number"),
    ("project_name", "Project Name"),
    ("task", "Task Name"),
    ("period", "Period"),
    ("person_number", "Person Number"),
    ("name", "Name"),
    ("hours", "Qty"),
    ("labor_type", "Labor Type"),
    ("time_type", "Time Type"),
    ("rate", "Rate"),
    ("subject", "Subject"),
    ("sender", "Sender"),
    ("rate_updated_at", "Rate Updated At"),
    ("rate_updated_by", "Rate Updated By"),
]


def export_snapshot_rows_to_excel(rows, output_path):
    """
    Renders build_outgoing_snapshot()-shaped entry rows to a formatted
    .xlsx -- used by the SharePoint Update/View Current/Finalize channel
    (services/sharepoint_service.py) both for a device's own
    current_<device_id>.xlsx and for the transient merged sheet Finalize
    prints. Not recorded into export_history -- these are working files
    for that channel, not a real export (see SHAREPOINT_SYNC_SPEC.md).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Current Sheet"

    ws.append([label for _, label in SNAPSHOT_COLUMNS])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER

    for row in rows:
        ws.append([row.get(key) for key, _ in SNAPSHOT_COLUMNS])
        for cell in ws[ws.max_row]:
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER

    for i, (_key, label) in enumerate(SNAPSHOT_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(len(label) + 2, 12)

    wb.save(output_path)


if __name__ == "__main__":
    init_db()
    print("Database initialized at", DB_PATH)
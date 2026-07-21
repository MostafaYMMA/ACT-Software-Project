"""
File-based SharePoint/OneDrive sync channel -- backs the SharePoint
Update / View Current / Finalize buttons (see ui/Pages/History.py and
services/sync_service.py). Transport is a plain locally-synced folder
(shutil/os file IO), not the Graph API -- see SHAREPOINT_SYNC_SPEC.md for
the full design this module implements.

This module owns ALL file IO for the channel: resolving the folder,
reading/writing boundary.json, and reading/writing/merging each device's
current_<device_id>.xlsx. It never touches the timecard DB tables
directly -- callers (sync_service) pass it rows already pulled from
storage_service, and it hands rows back for storage_service to render.

This is the only cross-device sync channel in the app -- there used to
also be an email-based one; it was removed in favor of this one.
"""

import json
import os
import tempfile
from datetime import datetime

from openpyxl import load_workbook

from storage_service import get_device_id, SNAPSHOT_COLUMNS, _received_month

BOUNDARY_FILENAME = "boundary.json"


class SharePointFolderError(RuntimeError):
    """Raised when sharepoint_folder is unset or isn't a real, existing
    directory. Callers must surface this as a clear error and do nothing
    else -- never fall back to writing files somewhere else instead."""


def require_sharepoint_folder(folder):
    if not folder or not os.path.isdir(folder):
        raise SharePointFolderError(
            "The SharePoint folder isn't set or doesn't exist. Set a valid, "
            "locally-synced folder path in Settings before using Update, "
            "View Current, or Finalize."
        )
    return folder


def _boundary_path(folder):
    return os.path.join(folder, BOUNDARY_FILENAME)


def read_boundary_date(folder):
    """Returns the ONE shared boundary_date ('YYYY-MM-DD'), or None if
    boundary.json doesn't exist yet -- first ever run, treat the boundary
    as 'the beginning of time' and include everything (spec sec 5)."""
    path = _boundary_path(folder)
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("boundary_date")


def _atomic_write(folder, real_path, prefix, suffix, write_fn):
    """Writes via a temp file in the same folder + os.replace, so the
    OneDrive client (or another device) never observes a half-written
    file mid-sync."""
    fd, tmp_path = tempfile.mkstemp(dir=folder, prefix=prefix, suffix=suffix)
    os.close(fd)
    try:
        write_fn(tmp_path)
        os.replace(tmp_path, real_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    return real_path


def write_boundary(folder, boundary_date):
    """Overwrites the single shared boundary.json -- the one reset point
    every device's Update filters against (RULE 2). Only Finalize calls
    this; simultaneous finalizes on two devices are an accepted edge case
    (last writer wins, spec sec 5)."""
    payload = {
        "boundary_date": boundary_date,
        "finalized_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "finalized_by": get_device_id(),
    }

    def write(tmp_path):
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)

    return _atomic_write(folder, _boundary_path(folder), ".boundary_", ".tmp", write)


def _device_filename(device_id):
    return f"current_{device_id}.xlsx"


def write_device_sheet(folder, rows):
    """Renders THIS device's own current-period rows to
    current_<device_id>.xlsx and writes it into the SharePoint folder,
    OVERWRITING whatever was already there (RULE 1: Update rebuilds from
    the DB every time, never appends -- clicking it five times in a row
    must produce the identical file)."""
    from storage_service import export_snapshot_rows_to_excel  # local import: avoids a hard circular import at module load

    device_id = get_device_id()
    dest_path = os.path.join(folder, _device_filename(device_id))
    _atomic_write(folder, dest_path, ".current_", ".xlsx", lambda tmp: export_snapshot_rows_to_excel(rows, tmp))
    return dest_path


def _read_device_sheet(path):
    """Reads one current_<device>.xlsx back into entry-shaped dicts,
    using the same column order storage_service.export_snapshot_rows_to_excel
    wrote (SNAPSHOT_COLUMNS) -- keep the two in sync."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter, None)  # header row
        keys = [key for key, _label in SNAPSHOT_COLUMNS]
        rows = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            rows.append(dict(zip(keys, values)))
        return rows
    finally:
        wb.close()


def list_device_sheets(folder):
    """Every current_*.xlsx currently in the folder -- all devices,
    including this one's just-written file. Sorted by filename so
    merge_device_sheets' first-seen-wins dedup is deterministic run to
    run, not dependent on OS directory-listing order."""
    return sorted(
        os.path.join(folder, name)
        for name in os.listdir(folder)
        if name.startswith("current_") and name.endswith(".xlsx")
    )


def _dedup_key(row):
    """The system-wide row identity: (day, Project Number, Task Name,
    person_number, received_month) -- the exact tuple
    storage_service._save_row uses. Reused, not reinvented (spec sec 12)."""
    return (
        row.get("day"),
        row.get("project_code"),
        row.get("task"),
        row.get("person_number"),
        _received_month(row.get("received")),
    )


def _device_id_from_filename(filename):
    """current_<device_id>.xlsx -> <device_id>. The inverse of
    _device_filename()."""
    stem = filename[:-len(".xlsx")] if filename.endswith(".xlsx") else filename
    prefix = "current_"
    return stem[len(prefix):] if stem.startswith(prefix) else None


def merge_device_sheets(folder):
    """Reads every device's current_*.xlsx and merges+dedups them on the
    system identity tuple. Read-only -- writes nothing (View Current must
    never mutate a device file or the DB). On a duplicate the first row
    seen wins, devices visited in sorted-filename order (see
    list_device_sheets) -- deterministic, not "whichever happened to be
    read last".

    Each row is tagged with "_origin_device_id" (parsed from the
    filename it came from) -- not a real data column, just enough for a
    caller (the SharePoint View Current window, see ui/Pages/History.py)
    to tell which rows are THIS device's own and therefore editable
    (a local DB record exists to attach an edit to) versus another
    device's (read-only here -- no local record exists for them)."""
    seen = {}
    sources = []
    for path in list_device_sheets(folder):
        filename = os.path.basename(path)
        sources.append(filename)
        origin_device_id = _device_id_from_filename(filename)
        for row in _read_device_sheet(path):
            key = _dedup_key(row)
            if key not in seen:
                row = dict(row)
                row["_origin_device_id"] = origin_device_id
                seen[key] = row
    return list(seen.values()), sources


def print_workbook(path, printer=None):
    """Prints an xlsx literally via Excel COM (matches the app's existing
    COM-first posture -- see filter_service.py's win32com usage for
    Outlook). Deliberately raises on any failure rather than swallowing it: the
    caller (sync_service.sharepoint_finalize) must NOT advance the
    boundary or reset the sheet when this fails, so a failed print
    leaves the period open for a retry (spec sec 7)."""
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(path))
        try:
            if printer:
                excel.ActivePrinter = printer
            wb.PrintOut()
        finally:
            wb.Close(SaveChanges=False)
    finally:
        excel.Quit()

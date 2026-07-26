"""
Serializes/deserializes cross-device email sync payloads (see
storage_service.build_outgoing_snapshot / build_rate_update_payload, and
sync_service.finalize_month) to/from a real .xlsx attachment instead of a
JSON blob -- opening the attachment in Excel now shows the actual rows,
not opaque JSON. Used by outlook_service.send_sync_mail / scan_sync_mails.

Kept as a standalone module (no win32com import) so this format and
services/sync_service.py's business logic can be exercised without
Outlook or Excel installed.

Layout -- two sheets, present as the payload kind needs them:
  "Meta" -- every top-level scalar field, one key/value row each. A
            "finalize" payload's nested "snapshot" dict's own scalars are
            prefixed "snapshot_" so they don't collide with finalize's
            own fields of the same name.
  "Rows" -- the row list, for "snapshot" and "finalize" kinds (a
            finalize's closing snapshot's rows) -- written with the same
            column layout storage_service.SNAPSHOT_COLUMNS uses for the
            SharePoint channel's current-sheet files, so both channels'
            attachments read the same way.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from storage_service import SNAPSHOT_COLUMNS

_ROW_KEYS = [key for key, _label in SNAPSHOT_COLUMNS]

# Shared look for both sheets' header row -- without this every column was
# sized off the HEADER LABEL's length only, never the data actually written
# into it, so a long "Project Name"/"Subject"/etc value just overflowed or
# got visually truncated next to another filled cell: this is what a user
# reported as the attachment looking like "garbage" when opened in Excel.
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4A4A4A")
_MAX_COLUMN_WIDTH = 60


def _style_header_row(ws, last_column):
    for col in range(1, last_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _autosize_columns(ws, header_labels):
    """Widens each column to fit the longest value actually written into
    it (header included), not just the header label -- capped so one huge
    outlier (a long Subject line) can't blow the sheet out sideways."""
    widths = [len(label) for label in header_labels]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, value in enumerate(row):
            if value is None:
                continue
            widths[i] = max(widths[i], min(len(str(value)), _MAX_COLUMN_WIDTH))
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width + 2, _MAX_COLUMN_WIDTH)


def _write_meta_sheet(ws, meta_items):
    ws.append(["Key", "Value"])
    for key, value in meta_items:
        ws.append([key, value])
    _style_header_row(ws, last_column=2)
    _autosize_columns(ws, ["Key", "Value"])


def _read_meta_sheet(ws):
    meta = {}
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # header
    for row in rows_iter:
        if not row or row[0] is None:
            continue
        meta[row[0]] = row[1]
    return meta


def _write_rows_sheet(ws, rows):
    labels = [label for _key, label in SNAPSHOT_COLUMNS]
    ws.append(labels)
    for row in rows:
        ws.append([row.get(key) for key in _ROW_KEYS])
    _style_header_row(ws, last_column=len(labels))
    _autosize_columns(ws, labels)


def _read_rows_sheet(ws):
    rows = []
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # header
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        rows.append(dict(zip(_ROW_KEYS, values)))
    return rows


def write_payload_workbook(payload, kind, output_path):
    """Renders a sync payload dict (kind in "snapshot"/"rate"/"finalize")
    to a real .xlsx at output_path."""
    wb = Workbook()
    meta_ws = wb.active
    meta_ws.title = "Meta"

    meta_items = [("kind", kind)]
    rows = None

    if kind == "rate":
        natural_key = list(payload.get("natural_key") or [])
        natural_key += [None] * (5 - len(natural_key))
        day, project_number, task_name, person_number, received_month = natural_key[:5]
        meta_items += [
            ("device_id", payload.get("device_id")),
            ("status", payload.get("status")),
            ("day", day),
            ("project_number", project_number),
            ("task_name", task_name),
            ("person_number", person_number),
            ("received_month", received_month),
            ("rate", payload.get("rate")),
            ("rate_updated_at", payload.get("rate_updated_at")),
            ("rate_updated_by", payload.get("rate_updated_by")),
        ]

    elif kind == "snapshot":
        meta_items += [
            ("device_id", payload.get("device_id")),
            ("seq", payload.get("seq")),
            ("generated_at", payload.get("generated_at")),
            ("period_start", payload.get("period_start")),
        ]
        rows = payload.get("rows") or []

    elif kind == "finalize":
        snapshot = payload.get("snapshot") or {}
        meta_items += [
            ("device_id", payload.get("device_id")),
            ("export_filename", payload.get("export_filename")),
            ("period_start", payload.get("period_start")),
            ("period_end", payload.get("period_end")),
            ("snapshot_device_id", snapshot.get("device_id")),
            ("snapshot_seq", snapshot.get("seq")),
            ("snapshot_generated_at", snapshot.get("generated_at")),
            ("snapshot_period_start", snapshot.get("period_start")),
        ]
        rows = snapshot.get("rows") or []

    else:
        raise ValueError(f"Unknown sync payload kind: {kind!r}")

    _write_meta_sheet(meta_ws, meta_items)

    if rows is not None:
        # Inserted BEFORE Meta and made the active sheet -- otherwise
        # opening the attachment lands on the small kind/device_id/seq
        # summary and the actual timecard rows sit unnoticed on a second
        # tab nobody clicks (this is exactly what looked like "no records
        # were sent" -- they were there, just not what greeted the eye).
        rows_ws = wb.create_sheet("Rows", 0)
        _write_rows_sheet(rows_ws, rows)
        wb.active = wb.sheetnames.index("Rows")

    wb.save(output_path)


def read_payload_workbook(path, kind):
    """The counterpart to write_payload_workbook -- reconstructs the
    same payload dict shape build_outgoing_snapshot /
    build_rate_update_payload / sync_service.finalize_month produced, for
    apply_incoming_snapshot / apply_rate_update /
    record_finalize_from_other_device to consume exactly as before this
    format changed. `kind` is the one already parsed off the mail
    subject (see outlook_service._SUBJECT_PATTERN) -- trusted over
    whatever the workbook's own "Meta" kind cell says, since the subject
    is what filter_service/scan_sync_mails already gate on.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        meta = _read_meta_sheet(wb["Meta"])
        rows = _read_rows_sheet(wb["Rows"]) if "Rows" in wb.sheetnames else None

        if kind == "rate":
            return {
                "device_id": meta.get("device_id"),
                "status": meta.get("status"),
                "natural_key": [
                    meta.get("day"), meta.get("project_number"), meta.get("task_name"),
                    meta.get("person_number"), meta.get("received_month"),
                ],
                "rate": meta.get("rate"),
                "rate_updated_at": meta.get("rate_updated_at"),
                "rate_updated_by": meta.get("rate_updated_by"),
            }

        if kind == "snapshot":
            return {
                "device_id": meta.get("device_id"),
                "seq": meta.get("seq"),
                "generated_at": meta.get("generated_at"),
                "period_start": meta.get("period_start"),
                "rows": rows or [],
            }

        if kind == "finalize":
            return {
                "device_id": meta.get("device_id"),
                "export_filename": meta.get("export_filename"),
                "period_start": meta.get("period_start"),
                "period_end": meta.get("period_end"),
                "snapshot": {
                    "device_id": meta.get("snapshot_device_id"),
                    "seq": meta.get("snapshot_seq"),
                    "generated_at": meta.get("snapshot_generated_at"),
                    "period_start": meta.get("snapshot_period_start"),
                    "rows": rows or [],
                },
            }

        raise ValueError(f"Unknown sync payload kind: {kind!r}")
    finally:
        wb.close()
